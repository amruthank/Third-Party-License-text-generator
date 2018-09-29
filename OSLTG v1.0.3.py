import os
import sys
import xlrd
import openpyxl
import re
import json
import webbrowser
from bs4 import BeautifulSoup
import requests
from difflib import SequenceMatcher
from titlecase import titlecase



error = {
         "package err" : "Invalid Package", "sheet err": "Wrong Sheet Name!",
         "header err" : "Expected labels are missing!!\nPlease add Component Name and License label to your sheet!!",
         "hyperlink err" : "Missing license URL!!",
         "page err" : "Web page 404 error!"
         }


#Beginning of the Class Excel!!
class Excel:

    def __init__(self, uploaded_file, sheet_name):
        self.uploaded_file = uploaded_file
        self.sheet_name = sheet_name
        self.file_type = ""
        

    #Function to verify the file type.
    def is_excel(self):
        file_name, self.file_type = (self.uploaded_file).rsplit(".",1)
        
        if self.file_type not in ["xls", "xlsx", "XLSX"]:
            return (False,None,None)
    
        return (True,file_name, self.file_type)
    #End of the function is_excel.



    #Function to parse the excel sheet.
    def parse_excel_sheet(self):

        if self.file_type == "xlsx":
            excel = openpyxl.load_workbook(self.uploaded_file)
            try:
                sheet = excel[self.sheet_name]
            except Exception:
                return ("sheet err", None)
        else:
            excel = xlrd.open_workbook(self.uploaded_file)

            try:
                #TODO: Sheet title case
                sheet = excel.sheet_by_name(self.sheet_name)
            except Exception:
                return ("sheet err", None)


        return (None, sheet) 
    #End of the function parse_excel_sheet.



    #Function to find the rows and columns numbers of component, license and row values. 
    def find_labeled_numbers(self, sheet):
        
        component_col = None
        license_col = None
        content_row = None

        if self.file_type == "xlsx":
            content_list = [[row,col] for row in range(1, sheet.max_row+1) for col in range(1, sheet.max_column+1) if sheet.cell(row,col).value == "Component Name" \
                        or sheet.cell(row,col).value == "Packages"]
        else:
            content_list = [[row,col] for row in range(sheet.nrows) for col in range(sheet.ncols) if sheet.cell(row,col).value == "Component Name" \
                            or sheet.cell(row,col).value == "Packages"]
            
        if len(content_list) != 0 and len(content_list[0])==2:
            content_row = int(content_list[0][0]) + 1  #conetnts will be below the present row.
            component_col = int(content_list[0][1]) 

        if self.file_type == "xlsx":
            content_list = [col for row in range(1, sheet.max_row+1) for col in range(1, sheet.max_column+1) if sheet.cell(row,col).value == "license" or \
                        sheet.cell(row,col).value == "License"]
        else:
            
            content_list = [col for row in range(sheet.nrows) for col in range(sheet.ncols) if sheet.cell(row,col).value == "license" or \
                            sheet.cell(row,col).value == "License"]

        if len(content_list) > 0:
            license_col = int(content_list[0])
       
        if (component_col is None or license_col is None or content_row is None):
            return ("header err", None, None, None)
        
        return (None, component_col,license_col,content_row)        
    #End of the function find_labeled_numbers.

#End of the Class Excel!!
    



#TODO: If the copyrights are in the multi line?? (https://github.com/matplotlib/matplotlib/blob/v2.1.2/LICENSE/LICENSE)

#Function to capture all the copyright notice from the license file
def capture_copyright(tag, copy_right, copyright_list, is_copyright):

    re_copyright = re.compile(r'[Cc]opyright (\([Cc@]\)|©|\d+).*', re.IGNORECASE)
    re_junk_copyright = re.compile('[Cc]opyright.*\<YEAR\>|yyyy|year.*', re.IGNORECASE)

    if re_junk_copyright.search(tag.text) != None:
        return ("", copyright_list, False)
        
    if tag.text != " ":

        found_copyright_notice = re_copyright.search(tag.text)
            
        if found_copyright_notice != None and found_copyright_notice.group(0) not in copyright_list: #Re-test for multiple copyrights in the license file.
            
            if len(tag.text) > 100: #If the copyright and license texts are text?!
                is_lengthy_text = True
                lengthy_text = tag.text

                for line in lengthy_text.split("\n"):
                    copyright_notice_from_lengthy_text = re_copyright.search(line)
                    if copyright_notice_from_lengthy_text != None:
                        copy_right = copyright_notice_from_lengthy_text.group(0)
                            
            elif len(tag.text) > 30 and not (re_copyright.search(found_copyright_notice.group(0))): #For the plain copyright text from the license.
                copy_right = ""
            else:            
                copy_right = found_copyright_notice.group(0)
            
            #if tag.text not in copyright_list and tag.text not in ["Copyright","copyright"]:
            if copy_right != "" and re.sub('All Rights Reserved.',"", copy_right, flags=re.IGNORECASE) not in copyright_list:
                copyright_list.append(re.sub('All Rights Reserved.',"", copy_right, flags=re.IGNORECASE))
                
            is_copyright = True


    return (copy_right, copyright_list, is_copyright)
#End of the function capture_copyright function.  



########################################################################Beginning Of License Terms#################################################################


'''
    License dictionary : license_info
    license_info keys: License names
    license_info values :

                {
                        Version number : License terms
                }

    Point to be take care while appending a new license details:

        1. First key index should be kept as a generic license information example None, appache 2.0 etc

'''

license_info = {
    
        ("Apache", "Apache License") :

            {

                "2.0" :
                
"""
TERMS AND CONDITIONS FOR USE, REPRODUCTION, AND DISTRIBUTION \n

1. Definitions.

"License" shall mean the terms and conditions for use, reproduction, and distribution as defined by Sections 1 through 9 of this document.

"Licensor" shall mean the copyright owner or entity authorized by the copyright owner that is granting the License.

"Legal Entity" shall mean the union of the acting entity and all other entities that control, are controlled by, or are under common control with that \
entity. For the purposes of this definition, "control" means (i) the power, direct or indirect, to cause the direction or management of such entity, \
whether by contract or otherwise, or (ii) ownership of fifty percent (50%) or more of the outstanding shares, or (iii) beneficial ownership of such entity.

"You" (or "Your") shall mean an individual or Legal Entity exercising permissions granted by this License.

"Source" form shall mean the preferred form for making modifications, including but not limited to software source code, documentation source, and \
configuration files.

"Object" form shall mean any form resulting from mechanical transformation or translation of a Source form, including but not limited to compiled object \
code, generated documentation, and conversions to other media types.

"Work" shall mean the work of authorship, whether in Source or Object form, made available under the License, as indicated by a copyright notice that is \
included in or attached to the work (an example is provided in the Appendix below).

"Derivative Works" shall mean any work, whether in Source or Object form, that is based on (or derived from) the Work and for which the editorial revisions,\
annotations, elaborations, or other modifications represent, as a whole, an original work of authorship. For the purposes of this License, Derivative Works\
shall not include works that remain separable from, or merely link (or bind by name) to the interfaces of, the Work and Derivative Works thereof.

"Contribution" shall mean any work of authorship, including the original version of the Work and any modifications or additions to that Work or Derivative \
Works thereof, that is intentionally submitted to Licensor for inclusion in the Work by the copyright owner or by an individual or Legal Entity authorized \
to submit on behalf of the copyright owner. For the purposes of this definition, "submitted" means any form of electronic, verbal, or written communication \
sent to the Licensor or its representatives, including but not limited to communication on electronic mailing lists, source code control systems, and issue \
tracking systems that are managed by, or on behalf of, the Licensor for the purpose of discussing and improving the Work, but excluding communication that \
is conspicuously marked or otherwise designated in writing by the copyright owner as "Not a Contribution."

"Contributor" shall mean Licensor and any individual or Legal Entity on behalf of whom a Contribution has been received by Licensor and subsequently \
incorporated within the Work.

2. Grant of Copyright License. Subject to the terms and conditions of this License, each Contributor hereby grants to You a perpetual, worldwide, \
non-exclusive, no-charge, royalty-free, irrevocable copyright license to reproduce, prepare Derivative Works of, publicly display, publicly perform, \
sublicense, and distribute the Work and such Derivative Works in Source or Object form.

3. Grant of Patent License. Subject to the terms and conditions of this License, each Contributor hereby grants to You a perpetual, worldwide, \
non-exclusive, no-charge, royalty-free, irrevocable (except as stated in this section) patent license to make, have made, use, offer to sell, sell, \
import, and otherwise transfer the Work, where such license applies only to those patent claims licensable by such Contributor that are necessarily \
infringed by their Contribution(s) alone or by combination of their Contribution(s) with the Work to which such Contribution(s) was submitted. If You \
institute patent litigation against any entity (including a cross-claim or counterclaim in a lawsuit) alleging that the Work or a Contribution incorporated \
within the Work constitutes direct or contributory patent infringement, then any patent licenses granted to You under this License for that Work shall \
terminate as of the date such litigation is filed.

4. Redistribution. You may reproduce and distribute copies of the Work or Derivative Works thereof in any medium, with or without modifications, and in \
Source or Object form, provided that You meet the following conditions:

You must give any other recipients of the Work or Derivative Works a copy of this License; and

You must cause any modified files to carry prominent notices stating that You changed the files; and

You must retain, in the Source form of any Derivative Works that You distribute, all copyright, patent, trademark, and attribution notices from the Source \
form of the Work, excluding those notices that do not pertain to any part of the Derivative Works; and

If the Work includes a "NOTICE" text file as part of its distribution, then any Derivative Works that You distribute must include a readable copy of the \
attribution notices contained within such NOTICE file, excluding those notices that do not pertain to any part of the Derivative Works, in at least one of \
the following places: within a NOTICE text file distributed as part of the Derivative Works; within the Source form or documentation, if provided along \
with the Derivative Works; or, within a display generated by the Derivative Works, if and wherever such third-party notices normally appear. The contents \
of the NOTICE file are for informational purposes only and do not modify the License. You may add Your own attribution notices within Derivative Works that \
You distribute, alongside or as an addendum to the NOTICE text from the Work, provided that such additional attribution notices cannot be construed as \
modifying the License. 

You may add Your own copyright statement to Your modifications and may provide additional or different license terms and conditions for use, reproduction,\
or distribution of Your modifications, or for any such Derivative Works as a whole, provided Your use, reproduction, and distribution of the Work otherwise \
complies with the conditions stated in this License.

5. Submission of Contributions. Unless You explicitly state otherwise, any Contribution intentionally submitted for inclusion in the Work by You to the \
Licensor shall be under the terms and conditions of this License, without any additional terms or conditions. Notwithstanding the above, nothing herein \
shall supersede or modify the terms of any separate license agreement you may have executed with Licensor regarding such Contributions.

6. Trademarks. This License does not grant permission to use the trade names, trademarks, service marks, or product names of the Licensor, except as \
required for reasonable and customary use in describing the origin of the Work and reproducing the content of the NOTICE file.

7. Disclaimer of Warranty. Unless required by applicable law or agreed to in writing, Licensor provides the Work (and each Contributor provides its \
Contributions) on an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied, including, without limitation, any warranties \
or conditions of TITLE, NON-INFRINGEMENT, MERCHANTABILITY, or FITNESS FOR A PARTICULAR PURPOSE. You are solely responsible for determining the \
appropriateness of using or redistributing the Work and assume any risks associated with Your exercise of permissions under this License.

8. Limitation of Liability. In no event and under no legal theory, whether in tort (including negligence), contract, or otherwise, unless required by \
applicable law (such as deliberate and grossly negligent acts) or agreed to in writing, shall any Contributor be liable to You for damages, including \
any direct, indirect, special, incidental, or consequential damages of any character arising as a result of this License or out of the use or inability \
to use the Work (including but not limited to damages for loss of goodwill, work stoppage, computer failure or malfunction, or any and all other commercial \
damages or losses), even if such Contributor has been advised of the possibility of such damages.

9. Accepting Warranty or Additional Liability. While redistributing the Work or Derivative Works thereof, You may choose to offer, and charge a fee for, \
acceptance of support, warranty, indemnity, or other liability obligations and/or rights consistent with this License. However, in accepting such \
obligations, You may act only on Your own behalf and on Your sole responsibility, not on behalf of any other Contributor, and only if You agree to \
indemnify, defend, and hold each Contributor harmless for any liability incurred by, or claims asserted against, such Contributor by reason of your \
accepting any such warranty or additional liability.

END OF TERMS AND CONDITIONS
""",
                

                "1.1" :

"""
The Apache Software License, Version 1.1

Copyright (c) 2000 The Apache Software Foundation.  All rights reserved.

Redistribution and use in source and binary forms, with or without modification, are permitted provided that the following conditions are met:

1. Redistributions of source code must retain the above copyright notice, this list of conditions and the following disclaimer.

2. Redistributions in binary form must reproduce the above copyright notice, this list of conditions and the following disclaimer in the documentation \
and/or other materials provided with the distribution.

3. The end-user documentation included with the redistribution, if any, must include the following acknowledgment:

"This product includes software developed by the Apache Software Foundation (http://www.apache.org/)."

Alternately, this acknowledgment may appear in the software itself, if and wherever such third-party acknowledgments normally appear.

4. The names "Apache" and "Apache Software Foundation" must not be used to endorse or promote products derived from this software without prior written \
permission. For written permission, please contact apache@apache.org.

5. Products derived from this software may not be called "Apache", nor may "Apache" appear in their name, without prior written permission of the Apache \
Software Foundation.

THIS SOFTWARE IS PROVIDED ``AS IS'' AND ANY EXPRESSED OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED WARRANTIES OF MERCHANTABILITY AND \
FITNESS FOR A PARTICULAR PURPOSE ARE
DISCLAIMED.  IN NO EVENT SHALL THE APACHE SOFTWARE FOUNDATION OR ITS CONTRIBUTORS BE LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR \
CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT
LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY THEORY OF \
LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY,
OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.
"""
            },

        
    ("MIT", "MIT License") :

            {
                "None":
                
"""
Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the "Software"), \
to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense, \
and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, \
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER \
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS \
IN THE SOFTWARE.
"""
             
            },

        "BSD" :
        
            {
                "2":

"""
Redistribution and use in source and binary forms, with or without modification, are permitted provided that the following conditions are met:

1. Redistributions of source code must retain the above copyright notice, this list of conditions and the following disclaimer.

2. Redistributions in binary form must reproduce the above copyright notice, this list of conditions and the following disclaimer in the \
documentation and/or other materials provided with the distribution.

THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, \
THE IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS \
BE LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE \
GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT \
LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH \
DAMAGE.

""",
                
                "3":

"""
Redistribution and use in source and binary forms, with or without modification, are permitted provided that the following conditions are met:

1. Redistributions of source code must retain the above copyright notice, this list of conditions and the following disclaimer.

2. Redistributions in binary form must reproduce the above copyright notice, this list of conditions and the following disclaimer in the \
documentation and/or other materials provided with the distribution.

3. Neither the name of the copyright holder nor the names of its contributors may be used to endorse or promote products derived from this software \
without specific prior written permission.

THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, \
THE IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS \
BE LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE \
GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT \
LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH \
DAMAGE

"""
            },

        ("PSF", "MATPLOTLIB License") :

            {
                "None":

"""
License agreement for matplotlib versions 1.3.0 and later
=========================================================

1. This LICENSE AGREEMENT is between the Matplotlib Development Team ("MDT"), and the Individual or Organization ("Licensee") accessing and otherwise using \
matplotlib software in source or binary form and its associated documentation.

2. Subject to the terms and conditions of this License Agreement, MDT hereby grants Licensee a nonexclusive, royalty-free, world-wide license to reproduce, \
analyze, test, perform and/or display publicly, prepare derivative works, distribute, and otherwise use matplotlib alone or in any derivative version, \
provided, however, that MDT's License Agreement and MDT's notice of copyright, i.e., "Copyright (c) 2012- Matplotlib Development Team; All Rights Reserved" \
are retained in matplotlib  alone or in any derivative version prepared by Licensee.

3. In the event Licensee prepares a derivative work that is based on or incorporates matplotlib or any part thereof, and wants to make the derivative \
work available to others as provided herein, then Licensee hereby agrees to include in any such work a brief summary of the changes made to matplotlib .

4. MDT is making matplotlib available to Licensee on an "AS IS" basis.  MDT MAKES NO REPRESENTATIONS OR WARRANTIES, EXPRESS OR IMPLIED.  BY WAY OF \
EXAMPLE, BUT NOT LIMITATION, MDT MAKES NO AND DISCLAIMS ANY REPRESENTATION OR WARRANTY OF MERCHANTABILITY OR FITNESS FOR ANY PARTICULAR PURPOSE OR \
THAT THE USE OF MATPLOTLIB WILL NOT INFRINGE ANY THIRD PARTY RIGHTS.

5. MDT SHALL NOT BE LIABLE TO LICENSEE OR ANY OTHER USERS OF MATPLOTLIB FOR ANY INCIDENTAL, SPECIAL, OR CONSEQUENTIAL DAMAGES OR LOSS AS A RESULT OF \
MODIFYING, DISTRIBUTING, OR OTHERWISE USING MATPLOTLIB , OR ANY DERIVATIVE THEREOF, EVEN IF ADVISED OF THE POSSIBILITY THEREOF.

6. This License Agreement will automatically terminate upon a material breach of its terms and conditions.

7. Nothing in this License Agreement shall be deemed to create any relationship of agency, partnership, or joint venture between MDT and Licensee.  \
This License Agreement does not grant permission to use MDT trademarks or trade name in a trademark sense to endorse or promote products or services \
of Licensee, or any third party.

8. By copying, installing or otherwise using matplotlib , Licensee agrees to be bound by the terms and conditions of this License Agreement.

License agreement for matplotlib versions prior to 1.3.0
========================================================

1. This LICENSE AGREEMENT is between John D. Hunter ("JDH"), and the Individual or Organization ("Licensee") accessing and otherwise using matplotlib \
software in source or binary form and its associated documentation.

2. Subject to the terms and conditions of this License Agreement, JDH hereby grants Licensee a nonexclusive, royalty-free, world-wide license to reproduce,\
analyze, test, perform and/or display publicly, prepare derivative works, distribute, and otherwise use matplotlib alone or in any derivative version, \
provided, however, that JDH's License Agreement and JDH's notice of copyright, i.e., "Copyright (c) 2002-2011 John D. Hunter; All Rights Reserved" are \
retained in matplotlib  alone or in any derivative version prepared by Licensee.

3. In the event Licensee prepares a derivative work that is based on or incorporates matplotlib  or any part thereof, and wants to make the derivative work \
available to others as provided herein, then Licensee hereby agrees to include in any such work a brief summary of the changes made to matplotlib.

4. JDH is making matplotlib  available to Licensee on an "AS IS" basis.  JDH MAKES NO REPRESENTATIONS OR WARRANTIES, EXPRESS OR IMPLIED.  BY WAY OF \
EXAMPLE, BUT NOT LIMITATION, JDH MAKES NO AND DISCLAIMS ANY REPRESENTATION OR WARRANTY OF MERCHANTABILITY OR FITNESS FOR ANY PARTICULAR PURPOSE OR THAT \
THE USE OF MATPLOTLIB WILL NOT INFRINGE ANY THIRD PARTY RIGHTS.

5. JDH SHALL NOT BE LIABLE TO LICENSEE OR ANY OTHER USERS OF MATPLOTLIB FOR ANY INCIDENTAL, SPECIAL, OR CONSEQUENTIAL DAMAGES OR LOSS AS A RESULT OF \
MODIFYING, DISTRIBUTING, OR OTHERWISE USING MATPLOTLIB , OR ANY DERIVATIVE THEREOF, EVEN IF ADVISED OF THE POSSIBILITY THEREOF.

6. This License Agreement will automatically terminate upon a material breach of its terms and conditions.

7. Nothing in this License Agreement shall be deemed to create any relationship of agency, partnership, or joint venture between JDH and Licensee.  This \
License Agreement does not grant permission to use JDH trademarks or trade name in a trademark sense to endorse or promote products or services of Licensee, or any third party.

8. By copying, installing or otherwise using matplotlib, Licensee agrees to be bound by the terms and conditions of this License Agreement. 
"""

            },

        "OpenSSL Combined License" :
                {
                    "None" :
"""
The OpenSSL toolkit stays under a double license, i.e. both the conditions of the OpenSSL License and the original SSLeay license apply to the toolkit.
See below for the actual license texts. Actually both licenses are BSD-style Open Source licenses. In case of any license issues related to OpenSSL \
please contact openssl-core@openssl.org.

OpenSSL License
---------------

====================================================================
Copyright (c) 1998-2017 The OpenSSL Project.  All rights reserved.
 
Redistribution and use in source and binary forms, with or without modification, are permitted provided that the following conditions are met:

1. Redistributions of source code must retain the above copyright notice, this list of conditions and the following disclaimer. 
2. Redistributions in binary form must reproduce the above copyright notice, this list of conditions and the following disclaimer in the documentation \
and/or other materials provided with the distribution.
3. All advertising materials mentioning features or use of this software must display the following acknowledgment:
    "This product includes software developed by the OpenSSL Project for use in the OpenSSL Toolkit. (http://www.openssl.org/)"
4. The names "OpenSSL Toolkit" and "OpenSSL Project" must not be used to endorse or promote products derived from this software without prior written \
permission. For written permission, please contact openssl-core@openssl.org.
5. Products derived from this software may not be called "OpenSSL" nor may "OpenSSL" appear in their names without prior written permission of the OpenSSL \
Project.
6. Redistributions of any form whatsoever must retain the following acknowledgment:
    "This product includes software developed by the OpenSSL Project for use in the OpenSSL Toolkit (http://www.openssl.org/)"

    THIS SOFTWARE IS PROVIDED BY THE OpenSSL PROJECT ``AS IS'' AND ANY EXPRESSED OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED \
    WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE DISCLAIMED.  IN NO EVENT SHALL THE OpenSSL PROJECT OR ITS CONTRIBUTORS BE \
    LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE \
    GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT \
    LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH \
    DAMAGE.
    ====================================================================
    This product includes cryptographic software written by Eric Young (eay@cryptsoft.com).  This product includes software written by Tim Hudson \
    (tjh@cryptsoft.com).

Original SSLeay License
-----------------------

Copyright (C) 1995-1998 Eric Young (eay@cryptsoft.com) All rights reserved.
This package is an SSL implementation written by Eric Young (eay@cryptsoft.com). The implementation was written so as to conform with Netscapes SSL.

This library is free for commercial and non-commercial use as long as the following conditions are aheared to.  The following conditions apply to all code \
found in this distribution, be it the RC4, RSA, lhash, DES, etc., code; not just the SSL code.  The SSL documentation included with this distribution is \
covered by the same copyright terms except that the holder is Tim Hudson (tjh@cryptsoft.com).

Copyright remains Eric Young's, and as such any Copyright notices in the code are not to be removed. If this package is used in a product, Eric Young \
should be given attribution as the author of the parts of the library used. This can be in the form of a textual message at program startup or in \
documentation (online or textual) provided with the package.

Redistribution and use in source and binary forms, with or without modification, are permitted provided that the following conditions are met:
    1. Redistributions of source code must retain the copyright notice, this list of conditions and the following disclaimer.
    2. Redistributions in binary form must reproduce the above copyright notice, this list of conditions and the following disclaimer in the documentation \
    and/or other materials provided with the distribution.
    3. All advertising materials mentioning features or use of this software must display the following acknowledgement:
        "This product includes cryptographic software written by Eric Young (eay@cryptsoft.com)"
        The word 'cryptographic' can be left out if the rouines from the library being used are not cryptographic related :-).
    4. If you include any Windows specific code (or a derivative thereof) from  the apps directory (application code) you must include an acknowledgement:
        "This product includes software written by Tim Hudson (tjh@cryptsoft.com)"

    THIS SOFTWARE IS PROVIDED BY THE OpenSSL PROJECT ``AS IS'' AND ANY EXPRESSED OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED \
    WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE DISCLAIMED.  IN NO EVENT SHALL THE OpenSSL PROJECT OR ITS CONTRIBUTORS BE \
    LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE \
    GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT \
    LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH \
    DAMAGE.

    The licence and distribution terms for any publically available version or derivative of this code cannot be changed.  i.e. this code cannot simply be \
    copied and put under another distribution licence [including the GNU Public Licence.]

"""
                },
        ("Microsoft Public License", "MS-PL") :

            {
                "None":
"""

1. Definitions
The terms "reproduce," "reproduction," "derivative works," and "distribution" have the
same meaning here as under U.S. copyright law.
A "contribution" is the original software, or any additions or changes to the software.
A "contributor" is any person that distributes its contribution under this license.
"Licensed patents" are a contributor's patent claims that read directly on its contribution.

2. Grant of Rights
(A) Copyright Grant- Subject to the terms of this license, including the license conditions and limitations in section 3, each contributor grants you a \
non-exclusive, worldwide, royalty-free copyright license to reproduce its contribution, prepare derivative works of its contribution, and distribute its \
contribution or any derivative works that you create.
(B) Patent Grant- Subject to the terms of this license, including the license conditions and limitations in section 3, each contributor grants you a \
non-exclusive, worldwide, royalty-free license under its licensed patents to make, have made, use, sell, offer for sale, import, and/or otherwise dispose \
of its contribution in the software or derivative works of the contribution in the software.

3. Conditions and Limitations
(A) No Trademark License- This license does not grant you rights to use any contributors' name, logo, or trademarks.
(B) If you bring a patent claim against any contributor over patents that you claim are infringed by the software, your patent license from such \
contributor to the software ends automatically.
(C) If you distribute any portion of the software, you must retain all copyright, patent, trademark, and attribution notices that are present in the \
software.
(D) If you distribute any portion of the software in source code form, you may do so only under this license by including a complete copy of this license \
with your distribution. If you distribute any portion of the software in compiled or object code form, you may only do so under a license that complies \
with this license.
(E) The software is licensed "as-is." You bear the risk of using it. The contributors give no express warranties, guarantees or conditions. You may have \
additional consumer rights under your local laws which this license cannot change. To the extent permitted under your local laws, the contributors exclude \
the implied warranties of merchantability, fitness for a particular purpose and non-infringement.


""" 
            },

        ("Mozilla Public License", "Mozilla", "MPL") :

                {
                    "2.0" :
                            
"""
1. Definitions
1.1. “Contributor”
means each individual or legal entity that creates, contributes to the creation of, or owns Covered Software.

1.2. “Contributor Version”
means the combination of the Contributions of others (if any) used by a Contributor and that particular Contributor’s Contribution.

1.3. “Contribution”
means Covered Software of a particular Contributor.

1.4. “Covered Software”
means Source Code Form to which the initial Contributor has attached the notice in Exhibit A, the Executable Form of such Source Code Form, and \
Modifications of such Source Code Form, in each case including portions thereof.

1.5. “Incompatible With Secondary Licenses”
means

that the initial Contributor has attached the notice described in Exhibit B to the Covered Software; or

that the Covered Software was made available under the terms of version 1.1 or earlier of the License, but not also under the terms of a Secondary License.

1.6. “Executable Form”
means any form of the work other than Source Code Form.

1.7. “Larger Work”
means a work that combines Covered Software with other material, in a separate file or files, that is not Covered Software.

1.8. “License”
means this document.

1.9. “Licensable”
means having the right to grant, to the maximum extent possible, whether at the time of the initial grant or subsequently, any and all of the rights \
conveyed by this License.

1.10. “Modifications”
means any of the following:

any file in Source Code Form that results from an addition to, deletion from, or modification of the contents of Covered Software; or

any new file in Source Code Form that contains any Covered Software.

1.11. “Patent Claims” of a Contributor
means any patent claim(s), including without limitation, method, process, and apparatus claims, in any patent Licensable by such Contributor that would be \
infringed, but for the grant of the License, by the making, using, selling, offering for sale, having made, import, or transfer of either its Contributions \
or its Contributor Version.

1.12. “Secondary License”
means either the GNU General Public License, Version 2.0, the GNU Lesser General Public License, Version 2.1, the GNU Affero General Public License, \
Version 3.0, or any later versions of those licenses.

1.13. “Source Code Form”
means the form of the work preferred for making modifications.

1.14. “You” (or “Your”)
means an individual or a legal entity exercising rights under this License. For legal entities, “You” includes any entity that controls, is controlled by, \
or is under common control with You. For purposes of this definition, “control” means (a) the power, direct or indirect, to cause the direction or \
management of such entity, whether by contract or otherwise, or (b) ownership of more than fifty percent (50%) of the outstanding shares or beneficial \
ownership of such entity.

2. License Grants and Conditions
2.1. Grants
Each Contributor hereby grants You a world-wide, royalty-free, non-exclusive license:

under intellectual property rights (other than patent or trademark) Licensable by such Contributor to use, reproduce, make available, modify, display, \
perform, distribute, and otherwise exploit its Contributions, either on an unmodified basis, with Modifications, or as part of a Larger Work; and

under Patent Claims of such Contributor to make, use, sell, offer for sale, have made, import, and otherwise transfer either its Contributions or its \
Contributor Version.

2.2. Effective Date
The licenses granted in Section 2.1 with respect to any Contribution become effective for each Contribution on the date the Contributor first distributes \
such Contribution.

2.3. Limitations on Grant Scope
The licenses granted in this Section 2 are the only rights granted under this License. No additional rights or licenses will be implied from the \
distribution or licensing of Covered Software under this License. Notwithstanding Section 2.1(b) above, no patent license is granted by a Contributor:

for any code that a Contributor has removed from Covered Software; or

for infringements caused by: (i) Your and any other third party’s modifications of Covered Software, or (ii) the combination of its Contributions with \
other software (except as part of its Contributor Version); or

under Patent Claims infringed by Covered Software in the absence of its Contributions.

This License does not grant any rights in the trademarks, service marks, or logos of any Contributor (except as may be necessary to comply with the notice \
requirements in Section 3.4).

2.4. Subsequent Licenses
No Contributor makes additional grants as a result of Your choice to distribute the Covered Software under a subsequent version of this License \
(see Section 10.2) or under the terms of a Secondary License (if permitted under the terms of Section 3.3).

2.5. Representation
Each Contributor represents that the Contributor believes its Contributions are its original creation(s) or it has sufficient rights to grant the rights \
to its Contributions conveyed by this License.

2.6. Fair Use
This License is not intended to limit any rights You have under applicable copyright doctrines of fair use, fair dealing, or other equivalents.

2.7. Conditions
Sections 3.1, 3.2, 3.3, and 3.4 are conditions of the licenses granted in Section 2.1.

3. Responsibilities
3.1. Distribution of Source Form
All distribution of Covered Software in Source Code Form, including any Modifications that You create or to which You contribute, must be under the terms \
of this License. You must inform recipients that the Source Code Form of the Covered Software is governed by the terms of this License, and how they can \
obtain a copy of this License. You may not attempt to alter or restrict the recipients’ rights in the Source Code Form.

3.2. Distribution of Executable Form
If You distribute Covered Software in Executable Form then:

such Covered Software must also be made available in Source Code Form, as described in Section 3.1, and You must inform recipients of the Executable Form \
how they can obtain a copy of such Source Code Form by reasonable means in a timely manner, at a charge no more than the cost of distribution to the \
recipient; and

You may distribute such Executable Form under the terms of this License, or sublicense it under different terms, provided that the license for the \
Executable Form does not attempt to limit or alter the recipients’ rights in the Source Code Form under this License.

3.3. Distribution of a Larger Work
You may create and distribute a Larger Work under terms of Your choice, provided that You also comply with the requirements of this License for the Covered \
Software. If the Larger Work is a combination of Covered Software with a work governed by one or more Secondary Licenses, and the Covered Software is not \
Incompatible With Secondary Licenses, this License permits You to additionally distribute such Covered Software under the terms of such Secondary \
License(s), so that the recipient of the Larger Work may, at their option, further distribute the Covered Software under the terms of either this License \
or such Secondary License(s).

3.4. Notices
You may not remove or alter the substance of any license notices (including copyright notices, patent notices, disclaimers of warranty, or limitations of \
liability) contained within the Source Code Form of the Covered Software, except that You may alter any license notices to the extent required to remedy \
known factual inaccuracies.

3.5. Application of Additional Terms
You may choose to offer, and to charge a fee for, warranty, support, indemnity or liability obligations to one or more recipients of Covered Software. \
However, You may do so only on Your own behalf, and not on behalf of any Contributor. You must make it absolutely clear that any such warranty, support, \
indemnity, or liability obligation is offered by You alone, and You hereby agree to indemnify every Contributor for any liability incurred by such \
Contributor as a result of warranty, support, indemnity or liability terms You offer. You may include additional disclaimers of warranty and limitations of \
liability specific to any jurisdiction.

4. Inability to Comply Due to Statute or Regulation
If it is impossible for You to comply with any of the terms of this License with respect to some or all of the Covered Software due to statute, judicial \
order, or regulation then You must: (a) comply with the terms of this License to the maximum extent possible; and (b) describe the limitations and the code \
they affect. Such description must be placed in a text file included with all distributions of the Covered Software under this License. Except to the \
extent prohibited by statute or regulation, such description must be sufficiently detailed for a recipient of ordinary skill to be able to understand it.

5. Termination
5.1. The rights granted under this License will terminate automatically if You fail to comply with any of its terms. However, if You become compliant, then \
the rights granted under this License from a particular Contributor are reinstated (a) provisionally, unless and until such Contributor explicitly and \
finally terminates Your grants, and (b) on an ongoing basis, if such Contributor fails to notify You of the non-compliance by some reasonable means prior \
to 60 days after You have come back into compliance. Moreover, Your grants from a particular Contributor are reinstated on an ongoing basis if such \
Contributor notifies You of the non-compliance by some reasonable means, this is the first time You have received notice of non-compliance with this \
License from such Contributor, and You become compliant prior to 30 days after Your receipt of the notice.

5.2. If You initiate litigation against any entity by asserting a patent infringement claim (excluding declaratory judgment actions, counter-claims, and \
cross-claims) alleging that a Contributor Version directly or indirectly infringes any patent, then the rights granted to You by any and all Contributors \
for the Covered Software under Section 2.1 of this License shall terminate.

5.3. In the event of termination under Sections 5.1 or 5.2 above, all end user license agreements (excluding distributors and resellers) which have been \
validly granted by You or Your distributors under this License prior to termination shall survive termination.

6. Disclaimer of Warranty
Covered Software is provided under this License on an “as is” basis, without warranty of any kind, either expressed, implied, or statutory, including, \
without limitation, warranties that the Covered Software is free of defects, merchantable, fit for a particular purpose or non-infringing. The entire risk \
as to the quality and performance of the Covered Software is with You. Should any Covered Software prove defective in any respect, You (not any Contributor) \
assume the cost of any necessary servicing, repair, or correction. This disclaimer of warranty constitutes an essential part of this License. No use of any \
Covered Software is authorized under this License except under this disclaimer.

7. Limitation of Liability
Under no circumstances and under no legal theory, whether tort (including negligence), contract, or otherwise, shall any Contributor, or anyone who \
distributes Covered Software as permitted above, be liable to You for any direct, indirect, special, incidental, or consequential damages of any character \
including, without limitation, damages for lost profits, loss of goodwill, work stoppage, computer failure or malfunction, or any and all other commercial \
damages or losses, even if such party shall have been informed of the possibility of such damages. This limitation of liability shall not apply to \
liability for death or personal injury resulting from such party’s negligence to the extent applicable law prohibits such limitation. Some jurisdictions do \
not allow the exclusion or limitation of incidental or consequential damages, so this exclusion and limitation may not apply to You.

8. Litigation
Any litigation relating to this License may be brought only in the courts of a jurisdiction where the defendant maintains its principal place of business \
and such litigation shall be governed by laws of that jurisdiction, without reference to its conflict-of-law provisions. Nothing in this Section shall \
prevent a party’s ability to bring cross-claims or counter-claims.

9. Miscellaneous
This License represents the complete agreement concerning the subject matter hereof. If any provision of this License is held to be unenforceable, such \
provision shall be reformed only to the extent necessary to make it enforceable. Any law or regulation which provides that the language of a contract shall \
be construed against the drafter shall not be used to construe this License against a Contributor.

10. Versions of the License
10.1. New Versions
Mozilla Foundation is the license steward. Except as provided in Section 10.3, no one other than the license steward has the right to modify or publish \
new versions of this License. Each version will be given a distinguishing version number.

10.2. Effect of New Versions
You may distribute the Covered Software under the terms of the version of the License under which You originally received the Covered Software, or under \
the terms of any subsequent version published by the license steward.

10.3. Modified Versions
If you create software not governed by this License, and you want to create a new license for such software, you may create and use a modified version of \
this License if you rename the license and remove any references to the name of the license steward (except to note that such modified license differs \
from this License).

10.4. Distributing Source Code Form that is Incompatible With Secondary Licenses
If You choose to distribute Source Code Form that is Incompatible With Secondary Licenses under the terms of this version of the License, the notice \
described in Exhibit B of this License must be attached.

Exhibit A - Source Code Form License Notice
This Source Code Form is subject to the terms of the Mozilla Public License, v. 2.0. If a copy of the MPL was not distributed with this file, You can \
obtain one at https://mozilla.org/MPL/2.0/.

If it is not possible or desirable to put the notice in a particular file, then You may include the notice in a location (such as a LICENSE file in a \
relevant directory) where a recipient would be likely to look for such a notice.

You may add additional accurate notices of copyright ownership.

""",

                    "1.1" :
"""
1. Definitions.
1.0.1. "Commercial Use"
means distribution or otherwise making the Covered Code available to a third party.
1.1. "Contributor"
means each entity that creates or contributes to the creation of Modifications.
1.2. "Contributor Version"
means the combination of the Original Code, prior Modifications used by a Contributor, and the Modifications made by that particular Contributor.
1.3. "Covered Code"
means the Original Code or Modifications or the combination of the Original Code and Modifications, in each case including portions thereof.
1.4. "Electronic Distribution Mechanism"
means a mechanism generally accepted in the software development community for the electronic transfer of data.
1.5. "Executable"
means Covered Code in any form other than Source Code.
1.6. "Initial Developer"
means the individual or entity identified as the Initial Developer in the Source Code notice required by Exhibit A.
1.7. "Larger Work"
means a work which combines Covered Code or portions thereof with code not governed by the terms of this License.
1.8. "License"
means this document.
1.8.1. "Licensable"
means having the right to grant, to the maximum extent possible, whether at the time of the initial grant or subsequently acquired, any and all of the \
rights conveyed herein.
1.9. "Modifications"
means any addition to or deletion from the substance or structure of either the Original Code or any previous Modifications. When Covered Code is released \
as a series of files, a Modification is:

Any addition to or deletion from the contents of a file containing Original Code or previous Modifications.
Any new file that contains any part of the Original Code or previous Modifications.
1.10. "Original Code"
means Source Code of computer software code which is described in the Source Code notice required by Exhibit A as Original Code, and which, at the time of \
its release under this License is not already Covered Code governed by this License.
1.10.1. "Patent Claims"
means any patent claim(s), now owned or hereafter acquired, including without limitation, method, process, and apparatus claims, in any patent Licensable \
by grantor.
1.11. "Source Code"
means the preferred form of the Covered Code for making modifications to it, including all modules it contains, plus any associated interface definition \
files, scripts used to control compilation and installation of an Executable, or source code differential comparisons against either the Original Code or \
another well known, available Covered Code of the Contributor's choice. The Source Code can be in a compressed or archival form, provided the appropriate \
decompression or de-archiving software is widely available for no charge.
1.12. "You" (or "Your")
means an individual or a legal entity exercising rights under, and complying with all of the terms of, this License or a future version of this License \
issued under Section 6.1. For legal entities, "You" includes any entity which controls, is controlled by, or is under common control with You. For purposes \
of this definition, "control" means (a) the power, direct or indirect, to cause the direction or management of such entity, whether by contract or \
otherwise, or (b) ownership of more than fifty percent (50%) of the outstanding shares or beneficial ownership of such entity.
2. Source Code License.
2.1. The Initial Developer Grant.
The Initial Developer hereby grants You a world-wide, royalty-free, non-exclusive license, subject to third party intellectual property claims:

under intellectual property rights (other than patent or trademark) Licensable by Initial Developer to use, reproduce, modify, display, perform, sublicense \
and distribute the Original Code (or portions thereof) with or without Modifications, and/or as part of a Larger Work; and
under Patents Claims infringed by the making, using or selling of Original Code, to make, have made, use, practice, sell, and offer for sale, and/or \
otherwise dispose of the Original Code (or portions thereof).
the licenses granted in this Section 2.1 (a) and (b) are effective on the date Initial Developer first distributes Original Code under the terms of this \
License.
Notwithstanding Section 2.1 (b) above, no patent license is granted: 1) for code that You delete from the Original Code; 2) separate from the Original \
Code; or 3) for infringements caused by: i) the modification of the Original Code or ii) the combination of the Original Code with other software or \
devices.
2.2. Contributor Grant.
Subject to third party intellectual property claims, each Contributor hereby grants You a world-wide, royalty-free, non-exclusive license

under intellectual property rights (other than patent or trademark) Licensable by Contributor, to use, reproduce, modify, display, perform, sublicense and \
distribute the Modifications created by such Contributor (or portions thereof) either on an unmodified basis, with other Modifications, as Covered Code \
and/or as part of a Larger Work; and
under Patent Claims infringed by the making, using, or selling of Modifications made by that Contributor either alone and/or in combination with its \
Contributor Version (or portions of such combination), to make, use, sell, offer for sale, have made, and/or otherwise dispose of: 1) Modifications made by \
that Contributor (or portions thereof); and 2) the combination of Modifications made by that Contributor with its Contributor Version (or portions of such \
combination).
the licenses granted in Sections 2.2 (a) and 2.2 (b) are effective on the date Contributor first makes Commercial Use of the Covered Code.
Notwithstanding Section 2.2 (b) above, no patent license is granted: 1) for any code that Contributor has deleted from the Contributor Version; 2) separate \
from the Contributor Version; 3) for infringements caused by: i) third party modifications of Contributor Version or ii) the combination of Modifications \
made by that Contributor with other software (except as part of the Contributor Version) or other devices; or 4) under Patent Claims infringed by Covered \
Code in the absence of Modifications made by that Contributor.
3. Distribution Obligations.
3.1. Application of License.
The Modifications which You create or to which You contribute are governed by the terms of this License, including without limitation Section 2.2. The \
Source Code version of Covered Code may be distributed only under the terms of this License or a future version of this License released under Section 6.1, \
and You must include a copy of this License with every copy of the Source Code You distribute. You may not offer or impose any terms on any Source Code \
version that alters or restricts the applicable version of this License or the recipients' rights hereunder. However, You may include an additional \
document offering the additional rights described in Section 3.5.

3.2. Availability of Source Code.
Any Modification which You create or to which You contribute must be made available in Source Code form under the terms of this License either on the same \
media as an Executable version or via an accepted Electronic Distribution Mechanism to anyone to whom you made an Executable version available; and if made \
available via Electronic Distribution Mechanism, must remain available for at least twelve (12) months after the date it initially became available, or at \
least six (6) months after a subsequent version of that particular Modification has been made available to such recipients. You are responsible for ensuring that the Source Code version remains available even if the Electronic Distribution Mechanism is maintained by a third party.

3.3. Description of Modifications.
You must cause all Covered Code to which You contribute to contain a file documenting the changes You made to create that Covered Code and the date of any \
change. You must include a prominent statement that the Modification is derived, directly or indirectly, from Original Code provided by the Initial \
Developer and including the name of the Initial Developer in (a) the Source Code, and (b) in any notice in an Executable version or related documentation \
in which You describe the origin or ownership of the Covered Code.

3.4. Intellectual Property Matters
(a) Third Party Claims
If Contributor has knowledge that a license under a third party's intellectual property rights is required to exercise the rights granted by such \
Contributor under Sections 2.1 or 2.2, Contributor must include a text file with the Source Code distribution titled "LEGAL" which describes the claim \
and the party making the claim in sufficient detail that a recipient will know whom to contact. If Contributor obtains such knowledge after the \
Modification is made available as described in Section 3.2, Contributor shall promptly modify the LEGAL file in all copies Contributor makes available \
thereafter and shall take other steps (such as notifying appropriate mailing lists or newsgroups) reasonably calculated to inform those who received the \
Covered Code that new knowledge has been obtained.

(b) Contributor APIs
If Contributor's Modifications include an application programming interface and Contributor has knowledge of patent licenses which are reasonably \
necessary to implement that API, Contributor must also include this information in the LEGAL file.

(c) Representations.
Contributor represents that, except as disclosed pursuant to Section 3.4 (a) above, Contributor believes that Contributor's Modifications are Contributor's \
original creation(s) and/or Contributor has sufficient rights to grant the rights conveyed by this License.

3.5. Required Notices.
You must duplicate the notice in Exhibit A in each file of the Source Code. If it is not possible to put such notice in a particular Source Code file due \
to its structure, then You must include such notice in a location (such as a relevant directory) where a user would be likely to look for such a notice. \
If You created one or more Modification(s) You may add your name as a Contributor to the notice described in Exhibit A. You must also duplicate this \
License in any documentation for the Source Code where You describe recipients' rights or ownership rights relating to Covered Code. You may choose to \
offer, and to charge a fee for, warranty, support, indemnity or liability obligations to one or more recipients of Covered Code. However, You may do so \
only on Your own behalf, and not on behalf of the Initial Developer or any Contributor. You must make it absolutely clear than any such warranty, support, \
indemnity or liability obligation is offered by You alone, and You hereby agree to indemnify the Initial Developer and every Contributor for any liability \
incurred by the Initial Developer or such Contributor as a result of warranty, support, indemnity or liability terms You offer.

3.6. Distribution of Executable Versions.
You may distribute Covered Code in Executable form only if the requirements of Sections 3.1, 3.2, 3.3, 3.4 and 3.5 have been met for that Covered Code, and \
if You include a notice stating that the Source Code version of the Covered Code is available under the terms of this License, including a description of \
how and where You have fulfilled the obligations of Section 3.2. The notice must be conspicuously included in any notice in an Executable version, related \
documentation or collateral in which You describe recipients' rights relating to the Covered Code. You may distribute the Executable version of Covered \
Code or ownership rights under a license of Your choice, which may contain terms different from this License, provided that You are in compliance with the \
terms of this License and that the license for the Executable version does not attempt to limit or alter the recipient's rights in the Source Code version \
from the rights set forth in this License. If You distribute the Executable version under a different license You must make it absolutely clear that any \
terms which differ from this License are offered by You alone, not by the Initial Developer or any Contributor. You hereby agree to indemnify the Initial \
Developer and every Contributor for any liability incurred by the Initial Developer or such Contributor as a result of any such terms You offer.

3.7. Larger Works.
You may create a Larger Work by combining Covered Code with other code not governed by the terms of this License and distribute the Larger Work as a single \
product. In such a case, You must make sure the requirements of this License are fulfilled for the Covered Code.

4. Inability to Comply Due to Statute or Regulation.
If it is impossible for You to comply with any of the terms of this License with respect to some or all of the Covered Code due to statute, judicial order, \
or regulation then You must: (a) comply with the terms of this License to the maximum extent possible; and (b) describe the limitations and the code they \
affect. Such description must be included in the LEGAL file described in Section 3.4 and must be included with all distributions of the Source Code. Except \
to the extent prohibited by statute or regulation, such description must be sufficiently detailed for a recipient of ordinary skill to be able to \
understand it.

5. Application of this License.
This License applies to code to which the Initial Developer has attached the notice in Exhibit A and to related Covered Code.

6. Versions of the License.
6.1. New Versions
Netscape Communications Corporation ("Netscape") may publish revised and/or new versions of the License from time to time. Each version will be given a \
distinguishing version number.

6.2. Effect of New Versions
Once Covered Code has been published under a particular version of the License, You may always continue to use it under the terms of that version. You may \
also choose to use such Covered Code under the terms of any subsequent version of the License published by Netscape. No one other than Netscape has the \
right to modify the terms applicable to Covered Code created under this License.

6.3. Derivative Works
If You create or use a modified version of this License (which you may only do in order to apply it to code which is not already Covered Code governed by \
this License), You must (a) rename Your license so that the phrases "Mozilla", "MOZILLAPL", "MOZPL", "Netscape", "MPL", "NPL" or any confusingly similar \
phrase do not appear in your license (except to note that your license differs from this License) and (b) otherwise make it clear that Your version of the \
license contains terms which differ from the Mozilla Public License and Netscape Public License. (Filling in the name of the Initial Developer, Original \
Code or Contributor in the notice described in Exhibit A shall not of themselves be deemed to be modifications of this License.)

7. DISCLAIMER OF WARRANTY
COVERED CODE IS PROVIDED UNDER THIS LICENSE ON AN "AS IS" BASIS, WITHOUT WARRANTY OF ANY KIND, EITHER EXPRESSED OR IMPLIED, INCLUDING, WITHOUT LIMITATION, \
WARRANTIES THAT THE COVERED CODE IS FREE OF DEFECTS, MERCHANTABLE, FIT FOR A PARTICULAR PURPOSE OR NON-INFRINGING. THE ENTIRE RISK AS TO THE QUALITY AND \
PERFORMANCE OF THE COVERED CODE IS WITH YOU. SHOULD ANY COVERED CODE PROVE DEFECTIVE IN ANY RESPECT, YOU (NOT THE INITIAL DEVELOPER OR ANY OTHER \
CONTRIBUTOR) ASSUME THE COST OF ANY NECESSARY SERVICING, REPAIR OR CORRECTION. THIS DISCLAIMER OF WARRANTY CONSTITUTES AN ESSENTIAL PART OF THIS LICENSE. \
NO USE OF ANY COVERED CODE IS AUTHORIZED HEREUNDER EXCEPT UNDER THIS DISCLAIMER.

8. Termination
8.1. This License and the rights granted hereunder will terminate automatically if You fail to comply with terms herein and fail to cure such breach within \
30 days of becoming aware of the breach. All sublicenses to the Covered Code which are properly granted shall survive any termination of this License. \
Provisions which, by their nature, must remain in effect beyond the termination of this License shall survive.

8.2. If You initiate litigation by asserting a patent infringement claim (excluding declatory judgment actions) against Initial Developer or a Contributor \
(the Initial Developer or Contributor against whom You file such action is referred to as "Participant") alleging that:

such Participant's Contributor Version directly or indirectly infringes any patent, then any and all rights granted by such Participant to You under \
Sections 2.1 and/or 2.2 of this License shall, upon 60 days notice from Participant terminate prospectively, unless if within 60 days after receipt of \
notice You either: (i) agree in writing to pay Participant a mutually agreeable reasonable royalty for Your past and future use of Modifications made by \
such Participant, or (ii) withdraw Your litigation claim with respect to the Contributor Version against such Participant. If within 60 days of notice, a \
reasonable royalty and payment arrangement are not mutually agreed upon in writing by the parties or the litigation claim is not withdrawn, the rights \
granted by Participant to You under Sections 2.1 and/or 2.2 automatically terminate at the expiration of the 60 day notice period specified above.
any software, hardware, or device, other than such Participant's Contributor Version, directly or indirectly infringes any patent, then any rights granted \
to You by such Participant under Sections 2.1(b) and 2.2(b) are revoked effective as of the date You first made, used, sold, distributed, or had made, \
Modifications made by that Participant.
8.3. If You assert a patent infringement claim against Participant alleging that such Participant's Contributor Version directly or indirectly infringes \
any patent where such claim is resolved (such as by license or settlement) prior to the initiation of patent infringement litigation, then the reasonable \
value of the licenses granted by such Participant under Sections 2.1 or 2.2 shall be taken into account in determining the amount or value of any payment \
or license.

8.4. In the event of termination under Sections 8.1 or 8.2 above, all end user license agreements (excluding distributors and resellers) which have been \
validly granted by You or any distributor hereunder prior to termination shall survive termination.

9. LIMITATION OF LIABILITY
UNDER NO CIRCUMSTANCES AND UNDER NO LEGAL THEORY, WHETHER TORT (INCLUDING NEGLIGENCE), CONTRACT, OR OTHERWISE, SHALL YOU, THE INITIAL DEVELOPER, ANY OTHER \
CONTRIBUTOR, OR ANY DISTRIBUTOR OF COVERED CODE, OR ANY SUPPLIER OF ANY OF SUCH PARTIES, BE LIABLE TO ANY PERSON FOR ANY INDIRECT, SPECIAL, INCIDENTAL, OR \
CONSEQUENTIAL DAMAGES OF ANY CHARACTER INCLUDING, WITHOUT LIMITATION, DAMAGES FOR LOSS OF GOODWILL, WORK STOPPAGE, COMPUTER FAILURE OR MALFUNCTION, OR ANY \
AND ALL OTHER COMMERCIAL DAMAGES OR LOSSES, EVEN IF SUCH PARTY SHALL HAVE BEEN INFORMED OF THE POSSIBILITY OF SUCH DAMAGES. THIS LIMITATION OF LIABILITY \
SHALL NOT APPLY TO LIABILITY FOR DEATH OR PERSONAL INJURY RESULTING FROM SUCH PARTY'S NEGLIGENCE TO THE EXTENT APPLICABLE LAW PROHIBITS SUCH LIMITATION. \
SOME JURISDICTIONS DO NOT ALLOW THE EXCLUSION OR LIMITATION OF INCIDENTAL OR CONSEQUENTIAL DAMAGES, SO THIS EXCLUSION AND LIMITATION MAY NOT APPLY TO YOU.

10. U.S. government end users
The Covered Code is a "commercial item," as that term is defined in 48 C.F.R. 2.101 (Oct. 1995), consisting of "commercial computer software" and \
"commercial computer software documentation," as such terms are used in 48 C.F.R. 12.212 (Sept. 1995). Consistent with 48 C.F.R. 12.212 and 48 C.F.R. \
227.7202-1 through 227.7202-4 (June 1995), all U.S. Government End Users acquire Covered Code with only those rights set forth herein.

11. Miscellaneous
This License represents the complete agreement concerning subject matter hereof. If any provision of this License is held to be unenforceable, such \
provision shall be reformed only to the extent necessary to make it enforceable. This License shall be governed by California law provisions \
(except to the extent applicable law, if any, provides otherwise), excluding its conflict-of-law provisions. With respect to disputes in which at \
least one party is a citizen of, or an entity chartered or registered to do business in the United States of America, any litigation relating to this \
License shall be subject to the jurisdiction of the Federal Courts of the Northern District of California, with venue lying in Santa Clara County, \
California, with the losing party responsible for costs, including without limitation, court costs and reasonable attorneys' fees and expenses. \
The application of the United Nations Convention on Contracts for the International Sale of Goods is expressly excluded. Any law or regulation which \
provides that the language of a contract shall be construed against the drafter shall not apply to this License.

12. Responsibility for claims
As between Initial Developer and the Contributors, each party is responsible for claims and damages arising, directly or indirectly, out of its \
utilization of rights under this License and You agree to work with Initial Developer and Contributors to distribute such responsibility on an \
equitable basis. Nothing herein is intended or shall be deemed to constitute any admission of liability.

13. Multiple-licensed code
Initial Developer may designate portions of the Covered Code as "Multiple-Licensed". "Multiple-Licensed" means that the Initial Developer permits you to \
utilize portions of the Covered Code under Your choice of the MPL or the alternative licenses, if any, specified by the Initial Developer in the file \
described in Exhibit A.


"""
                    },

    ("AFL", "Academic Free License","AFL V2", "AFL 2") :
            {
                "2.0":
"""

This Academic Free License (the "License") applies to any original work of authorship (the "Original Work") whose owner (the "Licensor") has placed the \
following notice immediately following the copyright notice for the Original Work:

Licensed under the Academic Free License version 2.0

1) Grant of Copyright License. Licensor hereby grants You a world-wide, royalty-free, non-exclusive, perpetual, sublicenseable license to do the following:
a) to reproduce the Original Work in copies;
b) to prepare derivative works ("Derivative Works") based upon the Original Work;
c) to distribute copies of the Original Work and Derivative Works to the public;
d) to perform the Original Work publicly; and
e) to display the Original Work publicly.
2) Grant of Patent License. Licensor hereby grants You a world-wide, royalty-free, non-exclusive, perpetual, sublicenseable license, under patent claims \
owned or controlled by the Licensor that are embodied in the Original Work as furnished by the Licensor, to make, use, sell and offer for sale the Original \
Work and Derivative Works.
3) Grant of Source Code License. The term "Source Code" means the preferred form of the Original Work for making modifications to it and all available \
documentation describing how to modify the Original Work. Licensor hereby agrees to provide a machine-readable copy of the Source Code of the Original \
Work along with each copy of the Original Work that Licensor distributes. Licensor reserves the right to satisfy this obligation by placing a \
machine-readable copy of the Source Code in an information repository reasonably calculated to permit inexpensive and convenient access by You for as \
long as Licensor continues to distribute the Original Work, and by publishing the address of that information repository in a notice immediately \
following the copyright notice that applies to the Original Work.
4) Exclusions From License Grant. Neither the names of Licensor, nor the names of any contributors to the Original Work, nor any of their trademarks or \
service marks, may be used to endorse or promote products derived from this Original Work without express prior written permission of the Licensor. \
Nothing in this License shall be deemed to grant any rights to trademarks, copyrights, patents, trade secrets or any other intellectual property of \
Licensor except as expressly stated herein. No patent license is granted to make, use, sell or offer to sell embodiments of any patent claims other than \
the licensed claims defined in Section 2. No right is granted to the trademarks of Licensor even if such marks are included in the Original Work. Nothing \
in this License shall be interpreted to prohibit Licensor from licensing under different terms from this License any Original Work that Licensor otherwise \
would have a right to license.
5) This section intentionally omitted.
6) Attribution Rights. You must retain, in the Source Code of any Derivative Works that You create, all copyright, patent or trademark notices from the \
Source Code of the Original Work, as well as any notices of licensing and any descriptive text identified therein as an "Attribution Notice." You must \
cause the Source Code for any Derivative Works that You create to carry a prominent Attribution Notice reasonably calculated to inform recipients that You \
have modified the Original Work.
7) Warranty of Provenance and Disclaimer of Warranty. Licensor warrants that the copyright in and to the Original Work and the patent rights granted herein \
by Licensor are owned by the Licensor or are sublicensed to You under the terms of this License with the permission of the contributor(s) of those \
copyrights and patent rights. Except as expressly stated in the immediately proceeding sentence, the Original Work is provided under this License on an \
"AS IS" BASIS and WITHOUT WARRANTY, either express or implied, including, without limitation, the warranties of NON-INFRINGEMENT, MERCHANTABILITY or \
FITNESS FOR A PARTICULAR PURPOSE. THE ENTIRE RISK AS TO THE QUALITY OF THE ORIGINAL WORK IS WITH YOU. This DISCLAIMER OF WARRANTY constitutes an essential \
part of this License. No license to Original Work is granted hereunder except under this disclaimer.
8) Limitation of Liability. Under no circumstances and under no legal theory, whether in tort (including negligence), contract, or otherwise, shall the \
Licensor be liable to any person for any direct, indirect, special, incidental, or consequential damages of any character arising as a result of this \
License or the use of the Original Work including, without limitation, damages for loss of goodwill, work stoppage, computer failure or malfunction, \
or any and all other commercial damages or losses. This limitation of liability shall not apply to liability for death or personal injury resulting from \
Licensor's negligence to the extent applicable law prohibits such limitation. Some jurisdictions do not allow the exclusion or limitation of incidental or \
consequential damages, so this exclusion and limitation may not apply to You.
9) Acceptance and Termination. If You distribute copies of the Original Work or a Derivative Work, You must make a reasonable effort under the \
circumstances to obtain the express assent of recipients to the terms of this License. Nothing else but this License (or another written agreement between \
Licensor and You) grants You permission to create Derivative Works based upon the Original Work or to exercise any of the rights granted in Section 1 \
herein, and any attempt to do so except under the terms of this License (or another written agreement between Licensor and You) is expressly prohibited \
by U.S. copyright law, the equivalent laws of other countries, and by international treaty. Therefore, by exercising any of the rights granted to You in \
Section 1 herein, You indicate Your acceptance of this License and all of its terms and conditions.
10) Termination for Patent Action. This License shall terminate automatically and You may no longer exercise any of the rights granted to You by this \
License as of the date You commence an action, including a cross-claim or counterclaim, for patent infringement (i) against Licensor with respect to a \
patent applicable to software or (ii) against any entity with respect to a patent applicable to the Original Work (but excluding combinations of the \
Original Work with other software or hardware).
11) Jurisdiction, Venue and Governing Law. Any action or suit relating to this License may be brought only in the courts of a jurisdiction wherein the \
Licensor resides or in which Licensor conducts its primary business, and under the laws of that jurisdiction excluding its conflict-of-law provisions. \
The application of the United Nations Convention on Contracts for the International Sale of Goods is expressly excluded. Any use of the Original Work \
outside the scope of this License or after its termination shall be subject to the requirements and penalties of the U.S. Copyright Act, 17 U.S.C. \
¤ 101 et seq., the equivalent laws of other countries, and international treaty. This section shall survive the termination of this License.
12) Attorneys Fees. In any action to enforce the terms of this License or seeking damages relating thereto, the prevailing party shall be entitled to \
recover its costs and expenses, including, without limitation, reasonable attorneys' fees and costs incurred in connection with such action, including \
any appeal of such action. This section shall survive the termination of this License.
13) Miscellaneous. This License represents the complete agreement concerning the subject matter hereof. If any provision of this License is held to be\
unenforceable, such provision shall be reformed only to the extent necessary to make it enforceable.
14) Definition of "You" in This License. "You" throughout this License, whether in upper or lower case, means an individual or a legal entity exercising \
rights under, and complying with all of the terms of, this License. For legal entities, "You" includes any entity that controls, is controlled by, or is \
under common control with you. For purposes of this definition, "control" means (i) the power, direct or indirect, to cause the direction or management of \
such entity, whether by contract or otherwise, or (ii) ownership of fifty percent (50%) or more of the outstanding shares, or (iii) beneficial ownership \
of such entity.
15) Right to Use. You may use the Original Work in all ways not otherwise restricted or conditioned by this License or by law, and Licensor promises not \
to interfere with or be responsible for such uses by You.


""",
                "3.0":
"""
1) Grant of Copyright License. Licensor grants You a worldwide, royalty-free, non-exclusive, sublicensable license, for the duration of the copyright, \
to do the following:

a) to reproduce the Original Work in copies, either alone or as part of a collective work;

b) to translate, adapt, alter, transform, modify, or arrange the Original Work, thereby creating derivative works ("Derivative Works") based upon the \
Original Work;

c) to distribute or communicate copies of the Original Work and Derivative Works to the public, under any license of your choice that does not contradict \
the terms and conditions, including Licensor's reserved rights and remedies, in this Academic Free License;

d) to perform the Original Work publicly; and

e) to display the Original Work publicly.

2) Grant of Patent License. Licensor grants You a worldwide, royalty-free, non-exclusive, sublicensable license, under patent claims owned or controlled \
by the Licensor that are embodied in the Original Work as furnished by the Licensor, for the duration of the patents, to make, use, sell, offer for sale, \
have made, and import the Original Work and Derivative Works.

3) Grant of Source Code License. The term "Source Code" means the preferred form of the Original Work for making modifications to it and all available \
documentation describing how to modify the Original Work. Licensor agrees to provide a machine-readable copy of the Source Code of the Original Work along \
with each copy of the Original Work that Licensor distributes. Licensor reserves the right to satisfy this obligation by placing a machine-readable copy of \
the Source Code in an information repository reasonably calculated to permit inexpensive and convenient access by You for as long as Licensor continues to \
distribute the Original Work.

4) Exclusions From License Grant. Neither the names of Licensor, nor the names of any contributors to the Original Work, nor any of their trademarks or \
service marks, may be used to endorse or promote products derived from this Original Work without express prior permission of the Licensor. Except as \
expressly stated herein, nothing in this License grants any license to Licensor's trademarks, copyrights, patents, trade secrets or any other intellectual \
property. No patent license is granted to make, use, sell, offer for sale, have made, or import embodiments of any patent claims other than the licensed \
claims defined in Section 2. No license is granted to the trademarks of Licensor even if such marks are included in the Original Work. Nothing in this \
License shall be interpreted to prohibit Licensor from licensing under terms different from this License any Original Work that Licensor otherwise would \
have a right to license.

5) External Deployment. The term "External Deployment" means the use, distribution, or communication of the Original Work or Derivative Works in any way \
such that the Original Work or Derivative Works may be used by anyone other than You, whether those works are distributed or communicated to those persons \
or made available as an application intended for use over a network. As an express condition for the grants of license hereunder, You must treat any \
External Deployment by You of the Original Work or a Derivative Work as a distribution under section 1(c).

6) Attribution Rights. You must retain, in the Source Code of any Derivative Works that You create, all copyright, patent, or trademark notices from the \
Source Code of the Original Work, as well as any notices of licensing and any descriptive text identified therein as an "Attribution Notice." You must \
cause the Source Code for any Derivative Works that You create to carry a prominent Attribution Notice reasonably calculated to inform recipients that You \
have modified the Original Work.

7) Warranty of Provenance and Disclaimer of Warranty. Licensor warrants that the copyright in and to the Original Work and the patent rights granted herein \
by Licensor are owned by the Licensor or are sublicensed to You under the terms of this License with the permission of the contributor(s) of those \
copyrights and patent rights. Except as expressly stated in the immediately preceding sentence, the Original Work is provided under this License on an "AS \
IS" BASIS and WITHOUT WARRANTY, either express or implied, including, without limitation, the warranties of non-infringement, merchantability or fitness \
for a particular purpose. THE ENTIRE RISK AS TO THE QUALITY OF THE ORIGINAL WORK IS WITH YOU. This DISCLAIMER OF WARRANTY constitutes an essential part of \
this License. No license to the Original Work is granted by this License except under this disclaimer.

8) Limitation of Liability. Under no circumstances and under no legal theory, whether in tort (including negligence), contract, or otherwise, shall the \
Licensor be liable to anyone for any indirect, special, incidental, or consequential damages of any character arising as a result of this License or the \
use of the Original Work including, without limitation, damages for loss of goodwill, work stoppage, computer failure or malfunction, or any and all other \
commercial damages or losses. This limitation of liability shall not apply to the extent applicable law prohibits such limitation.

9) Acceptance and Termination. If, at any time, You expressly assented to this License, that assent indicates your clear and irrevocable acceptance of this \
License and all of its terms and conditions. If You distribute or communicate copies of the Original Work or a Derivative Work, You must make a reasonable \
effort under the circumstances to obtain the express assent of recipients to the terms of this License. This License conditions your rights to undertake \
the activities listed in Section 1, including your right to create Derivative Works based upon the Original Work, and doing so without honoring these \
terms and conditions is prohibited by copyright law and international treaty. Nothing in this License is intended to affect copyright exceptions and \
limitations (including "fair use" or "fair dealing"). This License shall terminate immediately and You may no longer exercise any of the rights granted to \
You by this License upon your failure to honor the conditions in Section 1(c).

10) Termination for Patent Action. This License shall terminate automatically and You may no longer exercise any of the rights granted to You by this \
License as of the date You commence an action, including a cross-claim or counterclaim, against Licensor or any licensee alleging that the Original Work \
infringes a patent. This termination provision shall not apply for an action alleging patent infringement by combinations of the Original Work with other \
software or hardware.

11) Jurisdiction, Venue and Governing Law. Any action or suit relating to this License may be brought only in the courts of a jurisdiction wherein the \
Licensor resides or in which Licensor conducts its primary business, and under the laws of that jurisdiction excluding its conflict-of-law provisions. \
The application of the United Nations Convention on Contracts for the International Sale of Goods is expressly excluded. Any use of the Original Work \
outside the scope of this License or after its termination shall be subject to the requirements and penalties of copyright or patent law in the \
appropriate jurisdiction. This section shall survive the termination of this License.

12) Attorneys' Fees. In any action to enforce the terms of this License or seeking damages relating thereto, the prevailing party shall be entitled to \
recover its costs and expenses, including, without limitation, reasonable attorneys' fees and costs incurred in connection with such action, including \
any appeal of such action. This section shall survive the termination of this License.

13) Miscellaneous. If any provision of this License is held to be unenforceable, such provision shall be reformed only to the extent necessary to make \
it enforceable.

14) Definition of "You" in This License. "You" throughout this License, whether in upper or lower case, means an individual or a legal entity exercising \
rights under, and complying with all of the terms of, this License. For legal entities, "You" includes any entity that controls, is controlled by, or is \
under common control with you. For purposes of this definition, "control" means (i) the power, direct or indirect, to cause the direction or management of \
such entity, whether by contract or otherwise, or (ii) ownership of fifty percent (50%) or more of the outstanding shares, or (iii) beneficial ownership \
of such entity.

15) Right to Use. You may use the Original Work in all ways not otherwise restricted or conditioned by this License or by law, and Licensor promises not \
to interfere with or be responsible for such uses by You.

16) Modification of This License. This License is Copyright © 2005 Lawrence Rosen. Permission is granted to copy, distribute, or communicate this License \
without modification. Nothing in this License permits You to modify this License as applied to the Original Work or to Derivative Works. However, You may \
modify the text of this License and copy, distribute or communicate your modified version (the "Modified License") and apply it to other original works of \
authorship subject to the following conditions: (i) You may not indicate in any way that your Modified License is the "Academic Free License" or "AFL" and \
you may not use those names in the name of your Modified License; (ii) You must replace the notice specified in the first paragraph above with the notice \
"Licensed under <insert your license name here>" or with a notice of your own that is not confusingly similar to the notice in this License; and (iii) You \
may not claim that your original works are open source software unless your Modified License has been approved by Open Source Initiative (OSI) and You \
comply with its license review and certification process.


"""
            },


     ("Zlib", "Zlib License"):
        {
            "None":

"""

This software is provided 'as-is', without any express or implied warranty. In no event will the authors be held liable for any damages arising from the \
use of this software.

Permission is granted to anyone to use this software for any purpose, including commercial applications, and to alter it and redistribute it freely, \
subject to the following restrictions:

    1. The origin of this software must not be misrepresented; you must not claim that you wrote the original software. If you use this software in a \
    product, an acknowledgment in the product documentation would be appreciated but is not required.
    2. Altered source versions must be plainly marked as such, and must not be misrepresented as being the original software.
    3. This notice may not be removed or altered from any source distribution.


"""

            
        },

    ("General Public License", "GNU GPL", "GPL", "GNU General Public License") :

            {

                "2.0":
                
"""
Copyright (C) 1989, 1991 Free Software Foundation, Inc.

59 Temple Place, Suite 330, Boston, MA 02111-1307 USA Everyone is permitted to copy and distribute verbatim copies of this license document, \
but changing it is not allowed. Preamble

The licenses for most software are designed to take away your freedom to share and change it. By contrast, the GNU General Public License is \
intended to guarantee your freedom to share and change free software--to make sure the software is free for all its users. This General Public \
License applies to most of the Free Software Foundation's software and to any other program whose authors commit to using it. (Some other Free \
Software Foundation software is covered by the GNU Library General Public License instead.) You can apply it to your programs, too.

When we speak of free software, we are referring to freedom, not price. Our General Public Licenses are designed to make sure that you have the \
freedom to distribute copies of free software (and charge for this service if you wish), that you receive source code or can get it if you want it, \
that you can change the software or use pieces of it in new free programs; and that you know you can do these things.

To protect your rights, we need to make restrictions that forbid anyone to deny you these rights or to ask you to surrender the rights. These \
restrictions translate to certain responsibilities for you if you distribute copies of the software, or if you modify it.

For example, if you distribute copies of such a program, whether gratis or for a fee, you must give the recipients all the rights that you have.\
You must make sure that they, too, receive or can get the source code. And you must show them these terms so they know their rights.

We protect your rights with two steps: (1) copyright the software, and (2) offer you this license which gives you legal permission to copy, \
distribute and/or modify the software.

Also, for each author's protection and ours, we want to make certain that everyone understands that there is no warranty for this free software. \
If the software is modified by someone else and passed on, we want its recipients to know that what they have is not the original, so that any \
problems introduced by others will not reflect on the original authors' reputations.

Finally, any free program is threatened constantly by software patents. We wish to avoid the danger that redistributors of a free program will \
individually obtain patent licenses, in effect making the program proprietary. To prevent this, we have made it clear that any patent must be \
licensed for everyone's free use or not licensed at all.

The precise terms and conditions for copying, distribution and modification follow.

TERMS AND CONDITIONS FOR COPYING, DISTRIBUTION AND MODIFICATION

This License applies to any program or other work which contains a notice placed by the copyright holder saying it may be distributed under the \
terms of this General Public License. The "Program", below, refers to any such program or work, and a "work based on the Program" means either the \
Program or any derivative work under copyright law: that is to say, a work containing the Program or a portion of it, either verbatim or with \
modifications and/or translated into another language. (Hereinafter, translation is included without limitation in the term "modification".) \
Each licensee is addressed as "you".

Activities other than copying, distribution and modification are not covered by this License; they are outside its scope. The act of running the \
Program is not restricted, and the output from the Program is covered only if its contents constitute a work based on the Program (independent of \
having been made by running the Program). Whether that is true depends on what the Program does.

You may copy and distribute verbatim copies of the Program's source code as you receive it, in any medium, provided that you conspicuously and \
appropriately publish on each copy an appropriate copyright notice and disclaimer of warranty; keep intact all the notices that refer to this \
License and to the absence of any warranty; and give any other recipients of the Program a copy of this License along with the Program.

You may charge a fee for the physical act of transferring a copy, and you may at your option offer warranty protection in exchange for a fee.

You may modify your copy or copies of the Program or any portion of it, thus forming a work based on the Program, and copy and distribute such \
modifications or work under the terms of Section 1 above, provided that you also meet all of these conditions:

You must cause the modified files to carry prominent notices stating that you changed the files and the date of any change.

You must cause any work that you distribute or publish, that in whole or in part contains or is derived from the Program or any part thereof, to \
be licensed as a whole at no charge to all third parties under the terms of this License.

If the modified program normally reads commands interactively when run, you must cause it, when started running for such interactive use in the \
most ordinary way, to print or display an announcement including an appropriate copyright notice and a notice that there is no warranty (or else, \
saying that you provide a warranty) and that users may redistribute the program under these conditions, and telling the user how to view a copy of \
this License. (Exception: if the Program itself is interactive but does not normally  such an announcement, your work based on the Program is \
not required to print an announcement.)

These requirements apply to the modified work as a whole. If identifiable sections of that work are not derived from the Program, and can be \
reasonably considered independent and separate works in themselves, then this License, and its terms, do not apply to those sections when you \
distribute them as separate works. But when you distribute the same sections as part of a whole which is a work based on the Program, the distribution \
of the whole must be on the terms of this License, whose permissions for other licensees extend to the entire whole, and thus to each and every part \
regardless of who wrote it.

Thus, it is not the intent of this section to claim rights or contest your rights to work written entirely by you; rather, the intent is to exercise \
the right to control the distribution of derivative or collective works based on the Program.

In addition, mere aggregation of another work not based on the Program with the Program (or with a work based on the Program) on a volume of a storage \
or distribution medium does not bring the other work under the scope of this License.

You may copy and distribute the Program (or a work based on it, under Section 2) in object code or executable form under the terms of Sections 1 and \
2 above provided that you also do one of the following:

Accompany it with the complete corresponding machine-readable source code, which must be distributed under the terms of Sections 1 and 2 above on a \
medium customarily used for software interchange; or,

Accompany it with a written offer, valid for at least three years, to give any third party, for a charge no more than your cost of physically performing \
source distribution, a complete machine-readable copy of the corresponding source code, to be distributed under the terms of Sections 1 and 2 above on a \
medium customarily used for software interchange; or,

Accompany it with the information you received as to the offer to distribute corresponding source code. (This alternative is allowed only for noncommercial\
distribution and only if you received the program in object code or executable form with such an offer, in accord with Subsection b above.)

The source code for a work means the preferred form of the work for making modifications to it. For an executable work, complete source code means all the \
source code for all modules it contains, plus any associated interface definition files, plus the scripts used to control compilation and installation of \
the executable. However, as a special exception, the source code distributed need not include anything that is normally distributed (in either source or \
binary form) with the major components (compiler, kernel, and so on) of the operating system on which the executable runs, unless that component itself \
accompanies the executable.

If distribution of executable or object code is made by offering access to copy from a designated place, then offering equivalent access to copy the source\
code from the same place counts as distribution of the source code, even though third parties are not compelled to copy the source along with the object \
code.

You may not copy, modify, sublicense, or distribute the Program except as expressly provided under this License. Any attempt otherwise to copy, modify, \
sublicense or distribute the Program is void, and will automatically terminate your rights under this License. However, parties who have received copies, \
or rights, from you under this License will not have their licenses terminated so long as such parties remain in full compliance.

You are not required to accept this License, since you have not signed it. However, nothing else grants you permission to modify or distribute the Program \
or its derivative works. These actions are prohibited by law if you do not accept this License. Therefore, by modifying or distributing the Program (or \
any work based on the Program), you indicate your acceptance of this License to do so, and all its terms and conditions for copying, distributing or \
modifying the Program or works based on it.

Each time you redistribute the Program (or any work based on the Program), the recipient automatically receives a license from the original licensor to \
copy, distribute or modify the Program subject to these terms and conditions. You may not impose any further restrictions on the recipients' exercise of \
the rights granted herein. You are not responsible for enforcing compliance by third parties to this License.

If, as a consequence of a court judgment or allegation of patent infringement or for any other reason (not limited to patent issues), conditions are \
imposed on you (whether by court order, agreement or otherwise) that contradict the conditions of this License, they do not excuse you from the conditions \
of this License. If you cannot distribute so as to satisfy simultaneously your obligations under this License and any other pertinent obligations, then as \
a consequence you may not distribute the Program at all. For example, if a patent license would not permit royalty-free redistribution of the Program by \
all those who receive copies directly or indirectly through you, then the only way you could satisfy both it and this License would be to refrain entirely\
from distribution of the Program.

If any portion of this section is held invalid or unenforceable under any particular circumstance, the balance of the section is intended to apply and th
section as a whole is intended to apply in other circumstances.

It is not the purpose of this section to induce you to infringe any patents or other property right claims or to contest validity of any such claims; this \
section has the sole purpose of protecting the integrity of the free software distribution system, which is implemented by public license practices. Many \
people have made generous contributions to the wide range of software distributed through that system in reliance on consistent application of that system;\
it is up to the author/donor to decide if he or she is willing to distribute software through any other system and a licensee cannot impose that choice.

This section is intended to make thoroughly clear what is believed to be a consequence of the rest of this License.

If the distribution and/or use of the Program is restricted in certain countries either by patents or by copyrighted interfaces, the original copyright \
holder who places the Program under this License may add an explicit geographical distribution limitation excluding those countries, so that distribution \
is permitted only in or among countries not thus excluded. In such case, this License incorporates the limitation as if written in the body of this License.

The Free Software Foundation may publish revised and/or new versions of the General Public License from time to time. Such new versions will be similar in \
spirit to the present version, but may differ in detail to address new problems or concerns.

Each version is given a distinguishing version number. If the Program specifies a version number of this License which applies to it and "any later \
version",
you have the option of following the terms and conditions either of that version or of any later version published by the Free Software Foundation. If the \
Program does not specify a version number of this License, you may choose any version ever published by the Free Software Foundation.

If you wish to incorporate parts of the Program into other free programs whose distribution conditions are different, write to the author to ask for \
permission. For software which is copyrighted by the Free Software Foundation, write to the Free Software Foundation; we sometimes make exceptions for \
this. Our decision will be guided by the two goals of preserving the free status of all derivatives of our free software and of promoting the sharing and \
reuse of software generally.

NO WARRANTY

BECAUSE THE PROGRAM IS LICENSED FREE OF CHARGE, THERE IS NO WARRANTY FOR THE PROGRAM, TO THE EXTENT PERMITTED BY APPLICABLE LAW. EXCEPT WHEN OTHERWISE \
STATED IN WRITING THE COPYRIGHT HOLDERS AND/OR OTHER PARTIES PROVIDE THE PROGRAM "AS IS" WITHOUT WARRANTY OF ANY KIND, EITHER EXPRESSED OR IMPLIED, \
INCLUDING, BUT NOT LIMITED TO, THE IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE. THE ENTIRE RISK AS TO THE QUALITY AND \
PERFORMANCE OF THE PROGRAM IS WITH YOU. SHOULD THE PROGRAM PROVE DEFECTIVE, YOU ASSUME THE COST OF ALL NECESSARY SERVICING, REPAIR OR CORRECTION.

IN NO EVENT UNLESS REQUIRED BY APPLICABLE LAW OR AGREED TO IN WRITING WILL ANY COPYRIGHT HOLDER, OR ANY OTHER PARTY WHO MAY MODIFY AND/OR REDISTRIBUTE \
THE PROGRAM AS PERMITTED ABOVE, BE LIABLE TO YOU FOR DAMAGES, INCLUDING ANY GENERAL, SPECIAL, INCIDENTAL OR CONSEQUENTIAL DAMAGES ARISING OUT OF THE USE \
OR INABILITY TO USE THE PROGRAM (INCLUDING BUT NOT LIMITED TO LOSS OF DATA OR DATA BEING RENDERED INACCURATE OR LOSSES SUSTAINED BY YOU OR THIRD PARTIES \
OR A FAILURE OF THE PROGRAM TO OPERATE WITH ANY OTHER PROGRAMS), EVEN IF SUCH HOLDER OR OTHER PARTY HAS BEEN ADVISED OF THE POSSIBILITY OF SUCH DAMAGES.

END OF TERMS AND CONDITIONS

""",

            "3.0":

"""
Copyright © 2007 Free Software Foundation, Inc. <https://fsf.org/>

Everyone is permitted to copy and distribute verbatim copies of this license document, but changing it is not allowed.

Preamble
The GNU General Public License is a free, copyleft license for software and other kinds of works.

The licenses for most software and other practical works are designed to take away your freedom to share and change the works. By contrast, the GNU \
General Public License is intended to guarantee your freedom to share and change all versions of a program--to make sure it remains free software for \
all its users. We, the Free Software Foundation, use the GNU General Public License for most of our software; it applies also to any other work released \
this way by its authors. You can apply it to your programs, too.

When we speak of free software, we are referring to freedom, not price. Our General Public Licenses are designed to make sure that you have the freedom to \
distribute copies of free software (and charge for them if you wish), that you receive source code or can get it if you want it, that you can change the \
software or use pieces of it in new free programs, and that you know you can do these things.

To protect your rights, we need to prevent others from denying you these rights or asking you to surrender the rights. Therefore, you have certain \
responsibilities if you distribute copies of the software, or if you modify it: responsibilities to respect the freedom of others.

For example, if you distribute copies of such a program, whether gratis or for a fee, you must pass on to the recipients the same freedoms that you \
received. You must make sure that they, too, receive or can get the source code. And you must show them these terms so they know their rights.

Developers that use the GNU GPL protect your rights with two steps: (1) assert copyright on the software, and (2) offer you this License giving you \
legal permission to copy, distribute and/or modify it.

For the developers' and authors' protection, the GPL clearly explains that there is no warranty for this free software. For both users' and authors' \
sake, the GPL requires that modified versions be marked as changed, so that their problems will not be attributed erroneously to authors of previous \
versions.

Some devices are designed to deny users access to install or run modified versions of the software inside them, although the manufacturer can do so. \
This is fundamentally incompatible with the aim of protecting users' freedom to change the software. The systematic pattern of such abuse occurs in the \
area of products for individuals to use, which is precisely where it is most unacceptable. Therefore, we have designed this version of the GPL to prohibit \
the practice for those products. If such problems arise substantially in other domains, we stand ready to extend this provision to those domains in future \
versions of the GPL, as needed to protect the freedom of users.

Finally, every program is threatened constantly by software patents. States should not allow patents to restrict development and use of software on \
general-purpose computers, but in those that do, we wish to avoid the special danger that patents applied to a free program could make it effectively \
proprietary. To prevent this, the GPL assures that patents cannot be used to render the program non-free.

The precise terms and conditions for copying, distribution and modification follow.

TERMS AND CONDITIONS
0. Definitions.
“This License” refers to version 3 of the GNU General Public License.

“Copyright” also means copyright-like laws that apply to other kinds of works, such as semiconductor masks.

“The Program” refers to any copyrightable work licensed under this License. Each licensee is addressed as “you”. “Licensees” and “recipients” may be \
individuals or organizations.

To “modify” a work means to copy from or adapt all or part of the work in a fashion requiring copyright permission, other than the making of an exact \
copy. The resulting work is called a “modified version” of the earlier work or a work “based on” the earlier work.

A “covered work” means either the unmodified Program or a work based on the Program.

To “propagate” a work means to do anything with it that, without permission, would make you directly or secondarily liable for infringement under \
applicable copyright law, except executing it on a computer or modifying a private copy. Propagation includes copying, distribution (with or without \
modification), making available to the public, and in some countries other activities as well.

To “convey” a work means any kind of propagation that enables other parties to make or receive copies. Mere interaction with a user through a computer \
network, with no transfer of a copy, is not conveying.

An interactive user interface displays “Appropriate Legal Notices” to the extent that it includes a convenient and prominently visible feature that (1) \
displays an appropriate copyright notice, and (2) tells the user that there is no warranty for the work (except to the extent that warranties are provided),\
that licensees may convey the work under this License, and how to view a copy of this License. If the interface presents a list of user commands or options, \
such as a menu, a prominent item in the list meets this criterion.

1. Source Code.
The “source code” for a work means the preferred form of the work for making modifications to it. “Object code” means any non-source form of a work.

A “Standard Interface” means an interface that either is an official standard defined by a recognized standards body, or, in the case of interfaces \
specified for a particular programming language, one that is widely used among developers working in that language.

The “System Libraries” of an executable work include anything, other than the work as a whole, that (a) is included in the normal form of packaging a Major \
Component, but which is not part of that Major Component, and (b) serves only to enable use of the work with that Major Component, or to implement a \
Standard Interface for which an implementation is available to the public in source code form. A “Major Component”, in this context, means a major \
essential component (kernel, window system, and so on) of the specific operating system (if any) on which the executable work runs, or a compiler used \
to produce the work, or an object code interpreter used to run it.

The “Corresponding Source” for a work in object code form means all the source code needed to generate, install, and (for an executable work) run the \
object code and to modify the work, including scripts to control those activities. However, it does not include the work's System Libraries, or \
general-purpose tools or generally available free programs which are used unmodified in performing those activities but which are not part of the work. \
For example, Corresponding Source includes interface definition files associated with source files for the work, and the source code for shared libraries \
and dynamically linked subprograms that the work is specifically designed to require, such as by intimate data communication or control flow between those \
subprograms and other parts of the work.

The Corresponding Source need not include anything that users can regenerate automatically from other parts of the Corresponding Source.

The Corresponding Source for a work in source code form is that same work.

2. Basic Permissions.
All rights granted under this License are granted for the term of copyright on the Program, and are irrevocable provided the stated conditions are met. \
This License explicitly affirms your unlimited permission to run the unmodified Program. The output from running a covered work is covered by this License \
only if the output, given its content, constitutes a covered work. This License acknowledges your rights of fair use or other equivalent, as provided by \
copyright law.

You may make, run and propagate covered works that you do not convey, without conditions so long as your license otherwise remains in force. You may convey \
covered works to others for the sole purpose of having them make modifications exclusively for you, or provide you with facilities for running those works, \
provided that you comply with the terms of this License in conveying all material for which you do not control copyright. Those thus making or running the \
covered works for you must do so exclusively on your behalf, under your direction and control, on terms that prohibit them from making any copies of your \
copyrighted material outside their relationship with you.

Conveying under any other circumstances is permitted solely under the conditions stated below. Sublicensing is not allowed; section 10 makes it unnecessary.

3. Protecting Users' Legal Rights From Anti-Circumvention Law.
No covered work shall be deemed part of an effective technological measure under any applicable law fulfilling obligations under article 11 of the WIPO \
copyright treaty adopted on 20 December 1996, or similar laws prohibiting or restricting circumvention of such measures.

When you convey a covered work, you waive any legal power to forbid circumvention of technological measures to the extent such circumvention is effected \
by exercising rights under this License with respect to the covered work, and you disclaim any intention to limit operation or modification of the work as \
a means of enforcing, against the work's users, your or third parties' legal rights to forbid circumvention of technological measures.

4. Conveying Verbatim Copies.
You may convey verbatim copies of the Program's source code as you receive it, in any medium, provided that you conspicuously and appropriately publish on \
each copy an appropriate copyright notice; keep intact all notices stating that this License and any non-permissive terms added in accord with section 7 \
apply to the code; keep intact all notices of the absence of any warranty; and give all recipients a copy of this License along with the Program.

You may charge any price or no price for each copy that you convey, and you may offer support or warranty protection for a fee.

5. Conveying Modified Source Versions.
You may convey a work based on the Program, or the modifications to produce it from the Program, in the form of source code under the terms of section 4, \
provided that you also meet all of these conditions:

a) The work must carry prominent notices stating that you modified it, and giving a relevant date.
b) The work must carry prominent notices stating that it is released under this License and any conditions added under section 7. This requirement modifies \
the requirement in section 4 to “keep intact all notices”.
c) You must license the entire work, as a whole, under this License to anyone who comes into possession of a copy. This License will therefore apply, along \
with any applicable section 7 additional terms, to the whole of the work, and all its parts, regardless of how they are packaged. This License gives no \
permission to license the work in any other way, but it does not invalidate such permission if you have separately received it.
d) If the work has interactive user interfaces, each must display Appropriate Legal Notices; however, if the Program has interactive interfaces that do \
not display Appropriate Legal Notices, your work need not make them do so.
A compilation of a covered work with other separate and independent works, which are not by their nature extensions of the covered work, and which are \
not combined with it such as to form a larger program, in or on a volume of a storage or distribution medium, is called an “aggregate” if the compilation \
and its resulting copyright are not used to limit the access or legal rights of the compilation's users beyond what the individual works permit. Inclusion \
of a covered work in an aggregate does not cause this License to apply to the other parts of the aggregate.

6. Conveying Non-Source Forms.
You may convey a covered work in object code form under the terms of sections 4 and 5, provided that you also convey the machine-readable Corresponding \
Source under the terms of this License, in one of these ways:

a) Convey the object code in, or embodied in, a physical product (including a physical distribution medium), accompanied by the Corresponding Source fixed \
on a durable physical medium customarily used for software interchange.
b) Convey the object code in, or embodied in, a physical product (including a physical distribution medium), accompanied by a written offer, valid for at \
least three years and valid for as long as you offer spare parts or customer support for that product model, to give anyone who possesses the object code \
either (1) a copy of the Corresponding Source for all the software in the product that is covered by this License, on a durable physical medium customarily \
used for software interchange, for a price no more than your reasonable cost of physically performing this conveying of source, or (2) access to copy the \
Corresponding Source from a network server at no charge.
c) Convey individual copies of the object code with a copy of the written offer to provide the Corresponding Source. This alternative is allowed only \
occasionally and noncommercially, and only if you received the object code with such an offer, in accord with subsection 6b.
d) Convey the object code by offering access from a designated place (gratis or for a charge), and offer equivalent access to the Corresponding Source in \
the same way through the same place at no further charge. You need not require recipients to copy the Corresponding Source along with the object code. If \
the place to copy the object code is a network server, the Corresponding Source may be on a different server (operated by you or a third party) that \
supports equivalent copying facilities, provided you maintain clear directions next to the object code saying where to find the Corresponding Source. \
Regardless of what server hosts the Corresponding Source, you remain obligated to ensure that it is available for as long as needed to satisfy these \
requirements.
e) Convey the object code using peer-to-peer transmission, provided you inform other peers where the object code and Corresponding Source of the work \
are being offered to the general public at no charge under subsection 6d.
A separable portion of the object code, whose source code is excluded from the Corresponding Source as a System Library, need not be included in conveying \
the object code work.

A “User Product” is either (1) a “consumer product”, which means any tangible personal property which is normally used for personal, family, or household \
purposes, or (2) anything designed or sold for incorporation into a dwelling. In determining whether a product is a consumer product, doubtful cases shall \
be resolved in favor of coverage. For a particular product received by a particular user, “normally used” refers to a typical or common use of that class \
of product, regardless of the status of the particular user or of the way in which the particular user actually uses, or expects or is expected to use, the \
product. A product is a consumer product regardless of whether the product has substantial commercial, industrial or non-consumer uses, unless such uses \
represent the only significant mode of use of the product.

“Installation Information” for a User Product means any methods, procedures, authorization keys, or other information required to install and execute \
modified versions of a covered work in that User Product from a modified version of its Corresponding Source. The information must suffice to ensure that \
the continued functioning of the modified object code is in no case prevented or interfered with solely because modification has been made.

If you convey an object code work under this section in, or with, or specifically for use in, a User Product, and the conveying occurs as part of a \
transaction in which the right of possession and use of the User Product is transferred to the recipient in perpetuity or for a fixed term (regardless \
of how the transaction is characterized), the Corresponding Source conveyed under this section must be accompanied by the Installation Information. But \
this requirement does not apply if neither you nor any third party retains the ability to install modified object code on the User Product (for example, \
the work has been installed in ROM).

The requirement to provide Installation Information does not include a requirement to continue to provide support service, warranty, or updates for a work \
that has been modified or installed by the recipient, or for the User Product in which it has been modified or installed. Access to a network may be denied \
when the modification itself materially and adversely affects the operation of the network or violates the rules and protocols for communication across the \
network.

Corresponding Source conveyed, and Installation Information provided, in accord with this section must be in a format that is publicly documented (and with \
an implementation available to the public in source code form), and must require no special password or key for unpacking, reading or copying.

7. Additional Terms.
“Additional permissions” are terms that supplement the terms of this License by making exceptions from one or more of its conditions. Additional \
permissions that are applicable to the entire Program shall be treated as though they were included in this License, to the extent that they are valid \
under applicable law. If additional permissions apply only to part of the Program, that part may be used separately under those permissions, but the entire \
Program remains governed by this License without regard to the additional permissions.

When you convey a copy of a covered work, you may at your option remove any additional permissions from that copy, or from any part of it. (Additional \
permissions may be written to require their own removal in certain cases when you modify the work.) You may place additional permissions on material, \
added by you to a covered work, for which you have or can give appropriate copyright permission.

Notwithstanding any other provision of this License, for material you add to a covered work, you may (if authorized by the copyright holders of that \
material) supplement the terms of this License with terms:

a) Disclaiming warranty or limiting liability differently from the terms of sections 15 and 16 of this License; or
b) Requiring preservation of specified reasonable legal notices or author attributions in that material or in the Appropriate Legal Notices displayed \
by works containing it; or
c) Prohibiting misrepresentation of the origin of that material, or requiring that modified versions of such material be marked in reasonable ways as \
different from the original version; or
d) Limiting the use for publicity purposes of names of licensors or authors of the material; or
e) Declining to grant rights under trademark law for use of some trade names, trademarks, or service marks; or
f) Requiring indemnification of licensors and authors of that material by anyone who conveys the material (or modified versions of it) with contractual \
assumptions of liability to the recipient, for any liability that these contractual assumptions directly impose on those licensors and authors.
All other non-permissive additional terms are considered “further restrictions” within the meaning of section 10. If the Program as you received it, or \
any part of it, contains a notice stating that it is governed by this License along with a term that is a further restriction, you may remove that term. \
If a license document contains a further restriction but permits relicensing or conveying under this License, you may add to a covered work material \
governed by the terms of that license document, provided that the further restriction does not survive such relicensing or conveying.

If you add terms to a covered work in accord with this section, you must place, in the relevant source files, a statement of the additional terms that \
apply to those files, or a notice indicating where to find the applicable terms.

Additional terms, permissive or non-permissive, may be stated in the form of a separately written license, or stated as exceptions; the above requirements \
apply either way.

8. Termination.
You may not propagate or modify a covered work except as expressly provided under this License. Any attempt otherwise to propagate or modify it is void, \
and will automatically terminate your rights under this License (including any patent licenses granted under the third paragraph of section 11).

However, if you cease all violation of this License, then your license from a particular copyright holder is reinstated (a) provisionally, unless and until \
the copyright holder explicitly and finally terminates your license, and (b) permanently, if the copyright holder fails to notify you of the violation by \
some reasonable means prior to 60 days after the cessation.

Moreover, your license from a particular copyright holder is reinstated permanently if the copyright holder notifies you of the violation by some \
reasonable means, this is the first time you have received notice of violation of this License (for any work) from that copyright holder, and you cure \
the violation prior to 30 days after your receipt of the notice.

Termination of your rights under this section does not terminate the licenses of parties who have received copies or rights from you under this License. \
If your rights have been terminated and not permanently reinstated, you do not qualify to receive new licenses for the same material under section 10.

9. Acceptance Not Required for Having Copies.
You are not required to accept this License in order to receive or run a copy of the Program. Ancillary propagation of a covered work occurring solely as \
a consequence of using peer-to-peer transmission to receive a copy likewise does not require acceptance. However, nothing other than this License grants \
you permission to propagate or modify any covered work. These actions infringe copyright if you do not accept this License. Therefore, by modifying or \
propagating a covered work, you indicate your acceptance of this License to do so.

10. Automatic Licensing of Downstream Recipients.
Each time you convey a covered work, the recipient automatically receives a license from the original licensors, to run, modify and propagate that work, \
subject to this License. You are not responsible for enforcing compliance by third parties with this License.

An “entity transaction” is a transaction transferring control of an organization, or substantially all assets of one, or subdividing an organization, or \
merging organizations. If propagation of a covered work results from an entity transaction, each party to that transaction who receives a copy of the work \
also receives whatever licenses to the work the party's predecessor in interest had or could give under the previous paragraph, plus a right to possession \
of the Corresponding Source of the work from the predecessor in interest, if the predecessor has it or can get it with reasonable efforts.

You may not impose any further restrictions on the exercise of the rights granted or affirmed under this License. For example, you may not impose a license \
fee, royalty, or other charge for exercise of rights granted under this License, and you may not initiate litigation (including a cross-claim or \
counterclaim in a lawsuit) alleging that any patent claim is infringed by making, using, selling, offering for sale, or importing the Program or \
any portion of it.

11. Patents.
A “contributor” is a copyright holder who authorizes use under this License of the Program or a work on which the Program is based. The work thus \
licensed is called the contributor's “contributor version”.

A contributor's “essential patent claims” are all patent claims owned or controlled by the contributor, whether already acquired or hereafter acquired, \
that would be infringed by some manner, permitted by this License, of making, using, or selling its contributor version, but do not include claims that \
would be infringed only as a consequence of further modification of the contributor version. For purposes of this definition, “control” includes the right \
to grant patent sublicenses in a manner consistent with the requirements of this License.

Each contributor grants you a non-exclusive, worldwide, royalty-free patent license under the contributor's essential patent claims, to make, use, sell, \
offer for sale, import and otherwise run, modify and propagate the contents of its contributor version.

In the following three paragraphs, a “patent license” is any express agreement or commitment, however denominated, not to enforce a patent (such as an \
express permission to practice a patent or covenant not to sue for patent infringement). To “grant” such a patent license to a party means to make such an \
agreement or commitment not to enforce a patent against the party.

If you convey a covered work, knowingly relying on a patent license, and the Corresponding Source of the work is not available for anyone to copy, free of \
charge and under the terms of this License, through a publicly available network server or other readily accessible means, then you must either (1) cause \
the Corresponding Source to be so available, or (2) arrange to deprive yourself of the benefit of the patent license for this particular work, or (3) \
arrange, in a manner consistent with the requirements of this License, to extend the patent license to downstream recipients. “Knowingly relying” means \
you have actual knowledge that, but for the patent license, your conveying the covered work in a country, or your recipient's use of the covered work in \
a country, would infringe one or more identifiable patents in that country that you have reason to believe are valid.

If, pursuant to or in connection with a single transaction or arrangement, you convey, or propagate by procuring conveyance of, a covered work, and grant \
a patent license to some of the parties receiving the covered work authorizing them to use, propagate, modify or convey a specific copy of the covered \
work, then the patent license you grant is automatically extended to all recipients of the covered work and works based on it.

A patent license is “discriminatory” if it does not include within the scope of its coverage, prohibits the exercise of, or is conditioned on the \
non-exercise of one or more of the rights that are specifically granted under this License. You may not convey a covered work if you are a party to \
an arrangement with a third party that is in the business of distributing software, under which you make payment to the third party based on the extent \
of your activity of conveying the work, and under which the third party grants, to any of the parties who would receive the covered work from you, a \
discriminatory patent license (a) in connection with copies of the covered work conveyed by you (or copies made from those copies), or (b) primarily \
for and in connection with specific products or compilations that contain the covered work, unless you entered into that arrangement, or that patent \
license was granted, prior to 28 March 2007.

Nothing in this License shall be construed as excluding or limiting any implied license or other defenses to infringement that may otherwise be available \
to you under applicable patent law.

12. No Surrender of Others' Freedom.
If conditions are imposed on you (whether by court order, agreement or otherwise) that contradict the conditions of this License, they do not excuse you \
from the conditions of this License. If you cannot convey a covered work so as to satisfy simultaneously your obligations under this License and any other \
pertinent obligations, then as a consequence you may not convey it at all. For example, if you agree to terms that obligate you to collect a royalty for \
further conveying from those to whom you convey the Program, the only way you could satisfy both those terms and this License would be to refrain entirely \
from conveying the Program.

13. Use with the GNU Affero General Public License.
Notwithstanding any other provision of this License, you have permission to link or combine any covered work with a work licensed under version 3 of the \
GNU Affero General Public License into a single combined work, and to convey the resulting work. The terms of this License will continue to apply to the \
part which is the covered work, but the special requirements of the GNU Affero General Public License, section 13, concerning interaction through a network \
will apply to the combination as such.

14. Revised Versions of this License.
The Free Software Foundation may publish revised and/or new versions of the GNU General Public License from time to time. Such new versions will be \
similar in spirit to the present version, but may differ in detail to address new problems or concerns.

Each version is given a distinguishing version number. If the Program specifies that a certain numbered version of the GNU General Public License “or any \
later version” applies to it, you have the option of following the terms and conditions either of that numbered version or of any later version published by \
the Free Software Foundation. If the Program does not specify a version number of the GNU General Public License, you may choose any version ever published \
by the Free Software Foundation.

If the Program specifies that a proxy can decide which future versions of the GNU General Public License can be used, that proxy's public statement of \
acceptance of a version permanently authorizes you to choose that version for the Program.

Later license versions may give you additional or different permissions. However, no additional obligations are imposed on any author or copyright holder \
as a result of your choosing to follow a later version.

15. Disclaimer of Warranty.
THERE IS NO WARRANTY FOR THE PROGRAM, TO THE EXTENT PERMITTED BY APPLICABLE LAW. EXCEPT WHEN OTHERWISE STATED IN WRITING THE COPYRIGHT HOLDERS AND/OR OTHER \
PARTIES PROVIDE THE PROGRAM “AS IS” WITHOUT WARRANTY OF ANY KIND, EITHER EXPRESSED OR IMPLIED, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED WARRANTIES OF \
MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE. THE ENTIRE RISK AS TO THE QUALITY AND PERFORMANCE OF THE PROGRAM IS WITH YOU. SHOULD THE PROGRAM \
PROVE DEFECTIVE, YOU ASSUME THE COST OF ALL NECESSARY SERVICING, REPAIR OR CORRECTION.

16. Limitation of Liability.
IN NO EVENT UNLESS REQUIRED BY APPLICABLE LAW OR AGREED TO IN WRITING WILL ANY COPYRIGHT HOLDER, OR ANY OTHER PARTY WHO MODIFIES AND/OR CONVEYS THE \
PROGRAM AS PERMITTED ABOVE, BE LIABLE TO YOU FOR DAMAGES, INCLUDING ANY GENERAL, SPECIAL, INCIDENTAL OR CONSEQUENTIAL DAMAGES ARISING OUT OF THE USE OR \
INABILITY TO USE THE PROGRAM (INCLUDING BUT NOT LIMITED TO LOSS OF DATA OR DATA BEING RENDERED INACCURATE OR LOSSES SUSTAINED BY YOU OR THIRD PARTIES OR A \
FAILURE OF THE PROGRAM TO OPERATE WITH ANY OTHER PROGRAMS), EVEN IF SUCH HOLDER OR OTHER PARTY HAS BEEN ADVISED OF THE POSSIBILITY OF SUCH DAMAGES.

17. Interpretation of Sections 15 and 16.
If the disclaimer of warranty and limitation of liability provided above cannot be given local legal effect according to their terms, reviewing courts \
shall apply local law that most closely approximates an absolute waiver of all civil liability in connection with the Program, unless a warranty or \
assumption of liability accompanies a copy of the Program in return for a fee.

END OF TERMS AND CONDITIONS

"""
            },

    ("Lesser General Public License", "GNU Lesser General Public License", "Library General Public License", "Lesser General Public", "LGPL"):

        {
            "2.1":

"""
Version 2.1, February 1999

Copyright (C) 1991, 1999 Free Software Foundation, Inc.
51 Franklin Street, Fifth Floor, Boston, MA  02110-1301  USA
Everyone is permitted to copy and distribute verbatim copies
of this license document, but changing it is not allowed.


The licenses for most software are designed to take away your freedom to share and change it. By contrast, the GNU General Public Licenses are intended to \
guarantee your freedom to share and change free software--to make sure the software is free for all its users.

This license, the Lesser General Public License, applies to some specially designated software packages--typically libraries--of the Free Software \
Foundation and other authors who decide to use it. You can use it too, but we suggest you first think carefully about whether this license or the ordinary \
General Public License is the better strategy to use in any particular case, based on the explanations below.

When we speak of free software, we are referring to freedom of use, not price. Our General Public Licenses are designed to make sure that you have the \
freedom to distribute copies of free software (and charge for this service if you wish); that you receive source code or can get it if you want it; that \
you can change the software and use pieces of it in new free programs; and that you are informed that you can do these things.

To protect your rights, we need to make restrictions that forbid distributors to deny you these rights or to ask you to surrender these rights. These \
restrictions translate to certain responsibilities for you if you distribute copies of the library or if you modify it.

For example, if you distribute copies of the library, whether gratis or for a fee, you must give the recipients all the rights that we gave you. You must \
make sure that they, too, receive or can get the source code. If you link other code with the library, you must provide complete object files to the \
recipients, so that they can relink them with the library after making changes to the library and recompiling it. And you must show them these terms so \
they know their rights.

We protect your rights with a two-step method: (1) we copyright the library, and (2) we offer you this license, which gives you legal permission to copy, \
distribute and/or modify the library.

To protect each distributor, we want to make it very clear that there is no warranty for the free library. Also, if the library is modified by someone \
else \
and passed on, the recipients should know that what they have is not the original version, so that the original author's reputation will not be affected by \
problems that might be introduced by others.

Finally, software patents pose a constant threat to the existence of any free program. We wish to make sure that a company cannot effectively restrict the \
users of a free program by obtaining a restrictive license from a patent holder. Therefore, we insist that any patent license obtained for a version of \
the library must be consistent with the full freedom of use specified in this license.

Most GNU software, including some libraries, is covered by the ordinary GNU General Public License. This license, the GNU Lesser General Public License, \
applies to certain designated libraries, and is quite different from the ordinary General Public License. We use this license for certain libraries in \
order to permit linking those libraries into non-free programs.

When a program is linked with a library, whether statically or using a shared library, the combination of the two is legally speaking a combined work, a \
derivative of the original library. The ordinary General Public License therefore permits such linking only if the entire combination fits its criteria \
of freedom. The Lesser General Public License permits more lax criteria for linking other code with the library.

We call this license the "Lesser" General Public License because it does Less to protect the user's freedom than the ordinary General Public License. It \
also provides other free software developers Less of an advantage over competing non-free programs. These disadvantages are the reason we use the ordinary \
General Public License for many libraries. However, the Lesser license provides advantages in certain special circumstances.

For example, on rare occasions, there may be a special need to encourage the widest possible use of a certain library, so that it becomes a de-facto \
standard. To achieve this, non-free programs must be allowed to use the library. A more frequent case is that a free library does the same job as widely \
used non-free libraries. In this case, there is little to gain by limiting the free library to free software only, so we use the Lesser General Public \
License.

In other cases, permission to use a particular library in non-free programs enables a greater number of people to use a large body of free software. \
For example, permission to use the GNU C Library in non-free programs enables many more people to use the whole GNU operating system, as well as its \
variant, the GNU/Linux operating system.

Although the Lesser General Public License is Less protective of the users' freedom, it does ensure that the user of a program that is linked with the \
Library has the freedom and the wherewithal to run that program using a modified version of the Library.

The precise terms and conditions for copying, distribution and modification follow. Pay close attention to the difference between a "work based on the \
library" and a "work that uses the library". The former contains code derived from the library, whereas the latter must be combined with the library in \
order to run.

TERMS AND CONDITIONS FOR COPYING, DISTRIBUTION AND MODIFICATION
0. This License Agreement applies to any software library or other program which contains a notice placed by the copyright holder or other authorized \
party saying it may be distributed under the terms of this Lesser General Public License (also called "this License"). Each licensee is addressed as "you".

A "library" means a collection of software functions and/or data prepared so as to be conveniently linked with application programs (which use some of \
those functions and data) to form executables.

The "Library", below, refers to any such software library or work which has been distributed under these terms. A "work based on the Library" means either \
the Library or any derivative work under copyright law: that is to say, a work containing the Library or a portion of it, either verbatim or with \
modifications and/or translated straightforwardly into another language. (Hereinafter, translation is included without limitation in the term \
"modification".)

"Source code" for a work means the preferred form of the work for making modifications to it. For a library, complete source code means all the source \
code for all modules it contains, plus any associated interface definition files, plus the scripts used to control compilation and installation of the \
library.

Activities other than copying, distribution and modification are not covered by this License; they are outside its scope. The act of running a program \
using the Library is not restricted, and output from such a program is covered only if its contents constitute a work based on the Library (independent \
of the use of the Library in a tool for writing it). Whether that is true depends on what the Library does and what the program that uses the Library does.

1. You may copy and distribute verbatim copies of the Library's complete source code as you receive it, in any medium, provided that you conspicuously \
and appropriately publish on each copy an appropriate copyright notice and disclaimer of warranty; keep intact all the notices that refer to this License \
and to the absence of any warranty; and distribute a copy of this License along with the Library.

You may charge a fee for the physical act of transferring a copy, and you may at your option offer warranty protection in exchange for a fee.

2. You may modify your copy or copies of the Library or any portion of it, thus forming a work based on the Library, and copy and distribute such \
modifications or work under the terms of Section 1 above, provided that you also meet all of these conditions:

a) The modified work must itself be a software library.
b) You must cause the files modified to carry prominent notices stating that you changed the files and the date of any change.
c) You must cause the whole of the work to be licensed at no charge to all third parties under the terms of this License.
d) If a facility in the modified Library refers to a function or a table of data to be supplied by an application program that uses the facility, \
other than as an argument passed when the facility is invoked, then you must make a good faith effort to ensure that, in the event an application does \
not supply such function or table, the facility still operates, and performs whatever part of its purpose remains meaningful.
(For example, a function in a library to compute square roots has a purpose that is entirely well-defined independent of the application. Therefore, \
Subsection 2d requires that any application-supplied function or table used by this function must be optional: if the application does not supply it, \
the square root function must still compute square roots.)

These requirements apply to the modified work as a whole. If identifiable sections of that work are not derived from the Library, and can be reasonably \
considered independent and separate works in themselves, then this License, and its terms, do not apply to those sections when you distribute them as \
separate works. But when you distribute the same sections as part of a whole which is a work based on the Library, the distribution of the whole must be \
on the terms of this License, whose permissions for other licensees extend to the entire whole, and thus to each and every part regardless of who wrote it.

Thus, it is not the intent of this section to claim rights or contest your rights to work written entirely by you; rather, the intent is to exercise the \
right to control the distribution of derivative or collective works based on the Library.

In addition, mere aggregation of another work not based on the Library with the Library (or with a work based on the Library) on a volume of a storage or \
distribution medium does not bring the other work under the scope of this License.

3. You may opt to apply the terms of the ordinary GNU General Public License instead of this License to a given copy of the Library. To do this, you must \
alter all the notices that refer to this License, so that they refer to the ordinary GNU General Public License, version 2, instead of to this License. \
(If a newer version than version 2 of the ordinary GNU General Public License has appeared, then you can specify that version instead if you wish.) Do \
not make any other change in these notices.

Once this change is made in a given copy, it is irreversible for that copy, so the ordinary GNU General Public License applies to all subsequent copies \
and derivative works made from that copy.

This option is useful when you wish to copy part of the code of the Library into a program that is not a library.

4. You may copy and distribute the Library (or a portion or derivative of it, under Section 2) in object code or executable form under the terms of \
Sections 1 and 2 above provided that you accompany it with the complete corresponding machine-readable source code, which must be distributed under the \
terms of Sections 1 and 2 above on a medium customarily used for software interchange.

If distribution of object code is made by offering access to copy from a designated place, then offering equivalent access to copy the source code from \
the same place satisfies the requirement to distribute the source code, even though third parties are not compelled to copy the source along with the \
object code.

5. A program that contains no derivative of any portion of the Library, but is designed to work with the Library by being compiled or linked with it, is \
called a "work that uses the Library". Such a work, in isolation, is not a derivative work of the Library, and therefore falls outside the scope of this \
License.

However, linking a "work that uses the Library" with the Library creates an executable that is a derivative of the Library (because it contains portions \
of the Library), rather than a "work that uses the library". The executable is therefore covered by this License. Section 6 states terms for distribution \
of such executables.

When a "work that uses the Library" uses material from a header file that is part of the Library, the object code for the work may be a derivative work of \
the Library even though the source code is not. Whether this is true is especially significant if the work can be linked without the Library, or if the \
work is itself a library. The threshold for this to be true is not precisely defined by law.

If such an object file uses only numerical parameters, data structure layouts and accessors, and small macros and small inline functions (ten lines or \
less in length), then the use of the object file is unrestricted, regardless of whether it is legally a derivative work. (Executables containing this \
object code plus portions of the Library will still fall under Section 6.)

Otherwise, if the work is a derivative of the Library, you may distribute the object code for the work under the terms of Section 6. Any executables \
containing that work also fall under Section 6, whether or not they are linked directly with the Library itself.

6. As an exception to the Sections above, you may also combine or link a "work that uses the Library" with the Library to produce a work containing \
portions of the Library, and distribute that work under terms of your choice, provided that the terms permit modification of the work for the customer's \
own use and reverse engineering for debugging such modifications.

You must give prominent notice with each copy of the work that the Library is used in it and that the Library and its use are covered by this License. \
You must supply a copy of this License. If the work during execution displays copyright notices, you must include the copyright notice for the Library \
among them, as well as a reference directing the user to the copy of this License. Also, you must do one of these things:

a) Accompany the work with the complete corresponding machine-readable source code for the Library including whatever changes were used in the work \
(which must be distributed under Sections 1 and 2 above); and, if the work is an executable linked with the Library, with the complete machine-readable \
"work that uses the Library", as object code and/or source code, so that the user can modify the Library and then relink to produce a modified executable \
containing the modified Library. (It is understood that the user who changes the contents of definitions files in the Library will not necessarily be able \
to recompile the application to use the modified definitions.)
b) Use a suitable shared library mechanism for linking with the Library. A suitable mechanism is one that (1) uses at run time a copy of the library \
already present on the user's computer system, rather than copying library functions into the executable, and (2) will operate properly with a modified \
version of the library, if the user installs one, as long as the modified version is interface-compatible with the version that the work was made with.
c) Accompany the work with a written offer, valid for at least three years, to give the same user the materials specified in Subsection 6a, above, for a \
charge no more than the cost of performing this distribution.
d) If distribution of the work is made by offering access to copy from a designated place, offer equivalent access to copy the above specified materials \
from the same place.
e) Verify that the user has already received a copy of these materials or that you have already sent this user a copy.
For an executable, the required form of the "work that uses the Library" must include any data and utility programs needed for reproducing the executable \
from it. However, as a special exception, the materials to be distributed need not include anything that is normally distributed (in either source or \
binary form) with the major components (compiler, kernel, and so on) of the operating system on which the executable runs, unless that component itself \
accompanies the executable.

It may happen that this requirement contradicts the license restrictions of other proprietary libraries that do not normally accompany the operating system. \
Such a contradiction means you cannot use both them and the Library together in an executable that you distribute.

7. You may place library facilities that are a work based on the Library side-by-side in a single library together with other library facilities not \
covered by this License, and distribute such a combined library, provided that the separate distribution of the work based on the Library and of the other \
library facilities is otherwise permitted, and provided that you do these two things:

a) Accompany the combined library with a copy of the same work based on the Library, uncombined with any other library facilities. This must be \
distributed under the terms of the Sections above.
b) Give prominent notice with the combined library of the fact that part of it is a work based on the Library, and explaining where to find the \
accompanying uncombined form of the same work.
8. You may not copy, modify, sublicense, link with, or distribute the Library except as expressly provided under this License. Any attempt otherwise to \
copy, modify, sublicense, link with, or distribute the Library is void, and will automatically terminate your rights under this License. However, parties \
who have received copies, or rights, from you under this License will not have their licenses terminated so long as such parties remain in full compliance.

9. You are not required to accept this License, since you have not signed it. However, nothing else grants you permission to modify or distribute the \
Library or its derivative works. These actions are prohibited by law if you do not accept this License. Therefore, by modifying or distributing the \
Library (or any work based on the Library), you indicate your acceptance of this License to do so, and all its terms and conditions for copying, distributing or modifying the Library or works based on it.

10. Each time you redistribute the Library (or any work based on the Library), the recipient automatically receives a license from the original \
licensor to copy, distribute, link with or modify the Library subject to these terms and conditions. You may not impose any further restrictions on the \
recipients' exercise of the rights granted herein. You are not responsible for enforcing compliance by third parties with this License.

11. If, as a consequence of a court judgment or allegation of patent infringement or for any other reason (not limited to patent issues), conditions are \
imposed on you (whether by court order, agreement or otherwise) that contradict the conditions of this License, they do not excuse you from the conditions \
of this License. If you cannot distribute so as to satisfy simultaneously your obligations under this License and any other pertinent obligations, then as \
a consequence you may not distribute the Library at all. For example, if a patent license would not permit royalty-free redistribution of the Library by \
all those who receive copies directly or indirectly through you, then the only way you could satisfy both it and this License would be to refrain entirely \
from distribution of the Library.

If any portion of this section is held invalid or unenforceable under any particular circumstance, the balance of the section is intended to apply, \
and the section as a whole is intended to apply in other circumstances.

It is not the purpose of this section to induce you to infringe any patents or other property right claims or to contest validity of any such claims; \
this section has the sole purpose of protecting the integrity of the free software distribution system which is implemented by public license practices. \
Many people have made generous contributions to the wide range of software distributed through that system in reliance on consistent application of that \
system; it is up to the author/donor to decide if he or she is willing to distribute software through any other system and a licensee cannot impose that \
choice.

This section is intended to make thoroughly clear what is believed to be a consequence of the rest of this License.

12. If the distribution and/or use of the Library is restricted in certain countries either by patents or by copyrighted interfaces, the original \
copyright holder who places the Library under this License may add an explicit geographical distribution limitation excluding those countries, so \
that distribution is permitted only in or among countries not thus excluded. In such case, this License incorporates the limitation as if written \
in the body of this License.

13. The Free Software Foundation may publish revised and/or new versions of the Lesser General Public License from time to time. Such new versions \
will be similar in spirit to the present version, but may differ in detail to address new problems or concerns.

Each version is given a distinguishing version number. If the Library specifies a version number of this License which applies to it and "any later \
version", you have the option of following the terms and conditions either of that version or of any later version published by the Free Software \
Foundation. If the Library does not specify a license version number, you may choose any version ever published by the Free Software Foundation.

14. If you wish to incorporate parts of the Library into other free programs whose distribution conditions are incompatible with these, write to the \
author to ask for permission. For software which is copyrighted by the Free Software Foundation, write to the Free Software Foundation; we sometimes make \
exceptions for this. Our decision will be guided by the two goals of preserving the free status of all derivatives of our free software and of promoting \
the sharing and reuse of software generally.

NO WARRANTY

15. BECAUSE THE LIBRARY IS LICENSED FREE OF CHARGE, THERE IS NO WARRANTY FOR THE LIBRARY, TO THE EXTENT PERMITTED BY APPLICABLE LAW. EXCEPT WHEN OTHERWISE \
STATED IN WRITING THE COPYRIGHT HOLDERS AND/OR OTHER PARTIES PROVIDE THE LIBRARY "AS IS" WITHOUT WARRANTY OF ANY KIND, EITHER EXPRESSED OR IMPLIED, \
INCLUDING, BUT NOT LIMITED TO, THE IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE. THE ENTIRE RISK AS TO THE QUALITY AND \
PERFORMANCE OF THE LIBRARY IS WITH YOU. SHOULD THE LIBRARY PROVE DEFECTIVE, YOU ASSUME THE COST OF ALL NECESSARY SERVICING, REPAIR OR CORRECTION.

16. IN NO EVENT UNLESS REQUIRED BY APPLICABLE LAW OR AGREED TO IN WRITING WILL ANY COPYRIGHT HOLDER, OR ANY OTHER PARTY WHO MAY MODIFY AND/OR REDISTRIBUTE \
THE LIBRARY AS PERMITTED ABOVE, BE LIABLE TO YOU FOR DAMAGES, INCLUDING ANY GENERAL, SPECIAL, INCIDENTAL OR CONSEQUENTIAL DAMAGES ARISING OUT OF THE USE \
OR INABILITY TO USE THE LIBRARY (INCLUDING BUT NOT LIMITED TO LOSS OF DATA OR DATA BEING RENDERED INACCURATE OR LOSSES SUSTAINED BY YOU OR THIRD PARTIES \
OR A FAILURE OF THE LIBRARY TO OPERATE WITH ANY OTHER SOFTWARE), EVEN IF SUCH HOLDER OR OTHER PARTY HAS BEEN ADVISED OF THE POSSIBILITY OF SUCH DAMAGES.

END OF TERMS AND CONDITIONS
""" ,
            
            "3.0":
            
"""
Copyright (C) 2007 Free Software Foundation, Inc. <https://fsf.org/>
Everyone is permitted to copy and distribute verbatim copies \
of this license document, but changing it is not allowed. 


This version of the GNU Lesser General Public License incorporates the terms and conditions of version 3 of the GNU General Public License, \
supplemented by the additional permissions listed below.

0. Additional Definitions.

As used herein, "this License" refers to version 3 of the GNU Lesser General Public License, and the "GNU GPL" refers to version 3 of the GNU \ 
General Public License.

"The Library" refers to a covered work governed by this License, other than an Application or a Combined Work as defined below. 

An "Application" is any work that makes use of an interface provided by the Library, but which is not otherwise based on the Library. \
Defining a subclass of a class defined by the Library is deemed a mode of using an interface provided by the Library.

A "Combined Work" is a work produced by combining or linking an Application with the Library.  The particular version of the Library \
with which the Combined Work was made is also called the "Linked Version".

The "Minimal Corresponding Source" for a Combined Work means the Corresponding Source for the Combined Work, excluding any source code \
for portions of the Combined Work that, considered in isolation, are based on the Application, and not on the Linked Version.

The "Corresponding Application Code" for a Combined Work means the object code and/or source code for the Application, including any data \
and utility programs needed for reproducing the Combined Work from the Application, but excluding the System Libraries of the Combined Work. 

1. Exception to Section 3 of the GNU GPL.

You may convey a covered work under sections 3 and 4 of this License without being bound by section 3 of the GNU GPL.

2. Conveying Modified Versions.

If you modify a copy of the Library, and, in your modifications, a facility refers to a function or data to be supplied by an Application \
that uses the facility (other than as an argument passed when the facility is invoked), then you may convey a copy of the modified \
version:

a) under this License, provided that you make a good faith effort to ensure that, in the event an Application does not supply the \
function or data, the facility still operates, and performs whatever part of its purpose remains meaningful, or 

b) under the GNU GPL, with none of the additional permissions of this License applicable to that copy.

3. Object Code Incorporating Material from Library Header Files. 

The object code form of an Application may incorporate material from a header file that is part of the Library.  You may convey such object \
code under terms of your choice, provided that, if the incorporated material is not limited to numerical parameters, data structure \
layouts and accessors, or small macros, inline functions and templates (ten or fewer lines in length), you do both of the following:

a) Give prominent notice with each copy of the object code that the Library is used in it and that the Library and its use are \
covered by this License.

b) Accompany the object code with a copy of the GNU GPL and this license document.

4. Combined Works.

You may convey a Combined Work under terms of your choice that, taken together, effectively do not restrict modification of the \
portions of the Library contained in the Combined Work and reverse engineering for debugging such modifications, if you also do each of the following:

a) Give prominent notice with each copy of the Combined Work that the Library is used in it and that the Library and its use are \
covered by this License.

b) Accompany the Combined Work with a copy of the GNU GPL and this license document.

c) For a Combined Work that displays copyright notices during execution, include the copyright notice for the Library among \
these notices, as well as a reference directing the user to the copies of the GNU GPL and this license document.

d) Do one of the following:

0) Convey the Minimal Corresponding Source under the terms of this License, and the Corresponding Application Code in a form \
suitable for, and under terms that permit, the user to recombine or relink the Application with a modified version of \
the Linked Version to produce a modified Combined Work, in the manner specified by section 6 of the GNU GPL for conveying Corresponding Source.

1) Use a suitable shared library mechanism for linking with the Library.  A suitable mechanism is one that (a) uses at run time \
a copy of the Library already present on the user's computer system, and (b) will operate properly with a modified version \
of the Library that is interface-compatible with the Linked Version.

e) Provide Installation Information, but only if you would otherwise be required to provide such information under section 6 of the \
GNU GPL, and only to the extent that such information is necessary to install and execute a modified version of the \
Combined Work produced by recombining or relinking the Application with a modified version of the Linked Version. (If \
you use option 4d0, the Installation Information must accompany the Minimal Corresponding Source and Corresponding Application \
Code. If you use option 4d1, you must provide the Installation Information in the manner specified by section 6 of the GNU GPL \
for conveying Corresponding Source.)

5. Combined Libraries.

You may place library facilities that are a work based on the Library side by side in a single library together with other library \
facilities that are not Applications and are not covered by this License, and convey such a combined library under terms of your \
choice, if you do both of the following: 

a) Accompany the combined library with a copy of the same work based on the Library, uncombined with any other library facilities, \
conveyed under the terms of this License.

b) Give prominent notice with the combined library that part of it is a work based on the Library, and explaining where to find the \
accompanying uncombined form of the same work.

6. Revised Versions of the GNU Lesser General Public License.

The Free Software Foundation may publish revised and/or new versions of the GNU Lesser General Public License from time to time. Such new \
versions will be similar in spirit to the present version, but may differ in detail to address new problems or concerns.

Each version is given a distinguishing version number. If the Library as you received it specifies that a certain numbered version \
of the GNU Lesser General Public License "or any later version" applies to it, you have the option of following the terms and \
conditions either of that published version or of any later version published by the Free Software Foundation. If the Library as you \
received it does not specify a version number of the GNU Lesser General Public License, you may choose any version of the GNU Lesser \
General Public License ever published by the Free Software Foundation. 

If the Library as you received it specifies that a proxy can decide whether future versions of the GNU Lesser General Public License shall \
apply, that proxy's public statement of acceptance of any version is permanent authorization for you to choose that version for the Library.
"""
            }
         
}




hash_table = {}
hash_table["0"] = "Public Domain"
hash_table["1"] = "Public License"
hash_table["2"] = "Free public"
hash_table["3"] = "MIT"
hash_table["4"] = "Apache"
hash_table["5"] = "BSD"
hash_table["6"] = "ISC"
hash_table["7"] = "MPL"
hash_table["8"] = "Mozilla"

gpl_hash_table = {}
gpl_hash_table[0] = "GPL"
gpl_hash_table[1] = "LGPL"
gpl_hash_table[2] = "AGPL"
gpl_hash_table[3] = "General Public"




########################################################################END Of License Terms#################################################################


#Function to create a dictionary.
def create_third_party_dictioanry_contents(title, third_party_contents):
    
    license_name = ""
    license_version = None
    version = None
    dual_list = ["dual license", "dual licensed", "dual licensed under", "dual", "multi license", "multi licensed", "multi licensed under"]

    name_split_license = re.split(r'license', title, flags=re.IGNORECASE)[0] #Split license from the license name.
    name_split_digit = name_split_license.split(r'\b\d+[.]\d\b')[0] #Split numbers from the license name.
    name_split_spechar  = re.sub('[^A-Za-z0-9.]+', ' ', name_split_digit) #Split special characters from license name.
    

    name = name_split_spechar.split(" ", 1)[0] #Fetch first word from license name.
    
    if "gnu" == name.lower():
        name = name_split_spechar

    elif len([d for d in dual_list if ((name.lower()).find(d))!=-1]) > 0: #Handle dual license.
        pre_dual, suf_dual, found_name = None, None, None
        
        try:
            pre_dual,suf_dual = re.split(r'/or|and|\/', title)
        except Exception:
            pass

        for index in gpl_hash_table: #Try to catch GPL related licenses.
            if pre_dual and (pre_dual.lower()).find(gpl_hash_table[index].lower()) != -1:
                pre_dual = None
                break
            elif suf_dual and (suf_dual.lower()).find(gpl_hash_table[index].lower()) != -1:
                suf_dual = None
                break
       
                
        if pre_dual:
            found_name = pre_dual
        elif suf_dual:
            found_name = suf_dual

        if found_name and (found_name.lower()).find("dual") != -1:
            f_name = found_name
            found_name = re.sub(r'(dual licensed under|dual licensed|dual license)','', f_name, flags=re.IGNORECASE)
        
        is_hash_found = False
        for index in hash_table:
            
            if found_name and found_name.lower() == hash_table[index].lower() :
                name = hash_table[index]
                is_hash_found = True
                break
            elif found_name and re.search("public", hash_table[index], re.IGNORECASE) != None and re.search("public", found_name, re.IGNORECASE) != None:

                if (found_name.lower() == "public license"  and (found_name.lower()).find("eclipse")!= -1 and (title.find("General") != -1 or title.find("GNU") != -1 or title.find("LGPL") != -1 \
                                                          or title.find("AGPL") != -1)):
                    name = hash_table[index]
                    is_hash_found = True
                    break
            else:
                continue


        if not is_hash_found and found_name:
            try:
                name = found_name.split(r'\b\d+[.]\d\b[0-9]')[0] #Split version number, if present.
            except Exception:
                name = found_name
    else:
        pass

                       
    version = re.findall(r'\d[.\d]*', title) #Fetch license version, if any!
                                
    if len(version)>0:
        version = version[0]
    else:
        version = "None"

    if len(third_party_contents) != 0:

        #Checking for license name from exsting license dictionary.
        for key in license_info:

            if isinstance(key, tuple) and any(val in name for val in key):
                name = key[0]
            
        for key in third_party_contents.keys():
            
            if name.find(key) != -1 or key.find(name) != -1:
                license_name = key
                
                for index in third_party_contents[key].keys():
                    if index == version:
                        license_version = index
                        break
                    
                #Append new version to existing license dictionary.
                if license_version == None:
                    license_version = version
                    third_party_contents["%s"%(license_name)]["%s"%license_version] = {}
                    third_party_contents["%s"%(license_name)]["%s"%license_version]["terms"] = ""
                    
                return (third_party_contents, license_name, license_version)

             
    license_name = name
    license_version = version

    #Create new dictionary content.    
    third_party_contents["%s"%(license_name)] = {}
    third_party_contents["%s"%(license_name)]["%s"%license_version] = {}
    third_party_contents["%s"%(license_name)]["%s"%license_version]["terms"] = ""
                
    return (third_party_contents, license_name, license_version)
#End of the function create_third_party_dictioanry_contents.




#Parse the licese terms from the internet!!    
def extract_license_terms_from_web(soup):

    strings = ""
    terms = ""
    
    web_contents = soup.stripped_strings
    
    for line in web_contents:
        strings += "\n" + line
    
    possible_content_dict = {}

    possible_content_dict = {
                    "possible_beginnings" :
                        {
                            "0" : ["Definitions", "Preamble", "Permission to use", "MICROSOFT .NET LIBRARY"], #Declare
                            "1" : ["rights reserved", "License agreement"],  #Ignore
                            "2" : ["License", "license", "Copyright (c)"]
                        },

                    "possible_endings" :
                        {
                            "0" : ["Copy lines", "Copy permalink", "View git blame", "GitHub, Inc", "Jump to Line", "END OF TERMS AND CONDITIONS", \
                                   "extent permitted by applicable law"]
                        },

                    "junk_values" :
                        {
                            "0" : ["Copy lines", "Copy permalink", "View git blame", "GitHub, Inc", "Jump to Line", "Opensource.org"]
                        }
                }
        
    is_copyright = False
    is_terms = False
    is_begin_terms = False

    str_split_dict = {}
    str_split_dict["0"] = strings.splitlines()
    str_split_dict["1"] = strings.split("\n")
    
    for index in str_split_dict:
        
        #For license terms with copyrights!
        for line in str_split_dict[index]:

            if line == ".":
                continue
            
            re_exp = re.compile('[Cc][Oo][Pp][Yy][Rr][Ii][Gg][Hh][Tt] (\(Cc@\)|[0-9]+).*')
            re_exp_junk = re.compile('[Cc][Oo][Pp][Yy][Rr][Ii][Gg][Hh][Tt] \<YEAR\>*')

            #---------------------Beginnning of GitHub----------------------------------------
            if re_exp.search(line) != None or re_exp_junk.search(line) != None:
                if terms == "" and is_copyright == False:
                    is_copyright = True
                    continue
            
            if any(val in line for val in possible_content_dict["possible_beginnings"]["0"]) == True:
                is_begin_terms == True
                terms += str(line)
            elif any(val in line for val in possible_content_dict["possible_beginnings"]["1"]) == True:
                is_begin_terms == True
            elif any(val in line for val in possible_content_dict["possible_beginnings"]["2"]) == True:
                is_begin_terms == True                
            elif any(val in line for val in possible_content_dict["junk_values"]["0"]) == True:
                break
            elif terms != "" and any(val in line for val in possible_content_dict["possible_endings"]["0"]) == True:
                terms += "\n"+str(line)
                break
            else:
                if is_copyright == True or is_begin_terms == True:
                    terms += "\n"+str(line)
                else:
                    continue
            #---------------------End of GitHub------------------------------------------------
      
        #------------------------Beginning of single string block----------------------------
        if len(terms) < 50:
            
            if line == ".":
                continue
            
            is_terms = False
            terms = ""
            
            for line in str_split_dict[index]:

                if re.match('(.*?!)[Ll][Ii][Cc][Ee][Nn][Cc][Ee]$', line) != None:
                    is_terms = True
                    continue
                elif any(val in line for val in possible_content_dict["possible_beginnings"]["0"]) == True:
                    terms += str(line)
                    is_begin_terms = True
                elif is_begin_terms == False and any(val in line for val in possible_content_dict["possible_beginnings"]["1"]) == True:
                    is_begin_terms = True
                elif is_begin_terms == False and any(val in line for val in possible_content_dict["possible_beginnings"]["2"]) == True:
                    is_begin_terms = True    
                elif any(val in line for val in possible_content_dict["junk_values"]["0"]) == True:
                    break
                elif terms != "" and any(val in line for val in possible_content_dict["possible_endings"]["0"]) == True:
                    terms += "\n"+str(line)
                    is_terms == True
                    break
                else:
                    if is_terms == True or is_begin_terms == True:
                        terms += "\n" + str(line)
                    else:
                        continue
        #------------------------End of single string block--------------------------------
        if len(terms) > 400:
            break
        
    return terms


#Function to generate the third party license text file.
def generate_third_party_license_text(file_type, uploaded_file, sheet_name, sheet, component_col, license_col, row_val, third_party_contents):

    #Itertae over each component in the Excel sheet.
    if file_type == "xlsx":
        lval = sheet.max_row + 1
    else:
        lval = sheet.nrows
      
    for row in range(row_val, lval):
        
        is_copyright = False
        copyright_list = []
        terms_and_conditions_text = ""

        if sheet.cell(row,component_col).value != None:
            print("Parsing the component %s %s it's license %s"%((sheet.cell(row,component_col).value),(sheet.cell(row,component_col+1).value),\
                                                                 (sheet.cell(row,license_col).value)))
            (third_party_contents, title, version) = create_third_party_dictioanry_contents((sheet.cell(row,license_col).value), third_party_contents)
            
            try: #TODO: empty hyper link.
               
                if file_type == "xlsx" and sheet.cell(row,license_col).hyperlink != None:
                    hyperlink = sheet.cell(row,license_col).hyperlink.target
                elif file_type != "xlsx":
                    try:
                        hyperlink = sheet.hyperlink_map.get((row,license_col))
                    except Exception:
                        third_party_contents[title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)] = {}
                        third_party_contents[title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)]["%s"%(sheet.cell(row,component_col+1).value)] = []
                        third_party_contents[title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)]["%s"%(sheet.cell(row,component_col+1).value)].append\
                                                                                                                                (error["hyperlink err"])
                        continue
                else:
                    third_party_contents[title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)] = {}
                    third_party_contents[title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)]["%s"%(sheet.cell(row,component_col+1).value)] = []
                    third_party_contents[title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)]["%s"%(sheet.cell(row,component_col+1).value)].append\
                                                                                                                            (error["hyperlink err"])
                    continue
                                    
                try:
                    if file_type == "xlsx":
                        url = hyperlink
                    else:
                        url = hyperlink.url_or_path

                except Exception:
                   
                    third_party_contents[title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)] = {}
                    third_party_contents[title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)]["%s"%(sheet.cell(row,component_col+1).value)] = []
                    third_party_contents[title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)]["%s"%(sheet.cell(row,component_col+1).value)].append\
                                                                                                                            (error["hyperlink err"])
                    continue
                
                else:

                    if file_type == "xlsx":
                        response = requests.get(hyperlink)
                    else:
                        response = requests.get(hyperlink.url_or_path)
        
                    if response.status_code == 404:
                        third_party_contents[title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)] = {}
                        third_party_contents[title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)]["%s"%(sheet.cell(row,component_col+1).value)] = []
                        third_party_contents[title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)]["%s"%(sheet.cell(row,component_col+1).value)].append\
                                                                                                                                    (error["page err"])
                        continue    

            #TODO: Show in a dialog box.
            except requests.ConnectionError as e:
                print("OOPS!! Connection Error. Make sure, you are connected to the internet. Technical Details given below.\n")
                print(str(e))
            except requests.Timeout as e:
                print("OOPS!! Timeout Error")
                print(str(e))
            except requests.RequestException as e:
                print("OOPS!! General Error")
                print(str(e))
            except KeyboardInterrupt:
                print("Someone closed the program")
            else:
                #Duplicate components name and version.
                if (sheet.cell(row,component_col).value) in third_party_contents[title]["%s"%version]:
                    
                    if (sheet.cell(row,component_col+1).value) == third_party_contents[title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)]:
                        pass
                    else:
                       third_party_contents[title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)]["%s"%(sheet.cell(row,component_col+1).value)] = [] 
                else:
                    third_party_contents["%s"%title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)] = {}
                    third_party_contents["%s"%title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)]["%s"%(sheet.cell(row,component_col+1).value)] = []


            f_type = url.split(".")[-1]
            
            #TODO: Fetch license from *.pom file
            is_pom_file = False
            
            if f_type in ["nuspec"]:
                print("It's a xml file!! please provide the correct license URL!!")
            elif f_type == "pom":
                is_pom_file = True
            elif f_type == "json":
                pass
            
            page = requests.get(url)
            
            soup = BeautifulSoup(page.content,'html.parser')
            
            #hyperlink.dump(header = 60 * '-')
            
            copy_rigth = ""

            lengthy_text = ""
            is_lengthy_text = False
            copy_rigth = ""

            
             
            #Capture copyright notice.
            for tag in soup():
                copy_rigth, copyright_list, is_copyright = capture_copyright(tag, copy_rigth, copyright_list, is_copyright)
                         
            
            #Extract license terms from internet!!
            try:
                web_terms = extract_license_terms_from_web(soup)
            except Exception:
                web_terms = ""
                pass
            
            if not is_lengthy_text:

                #Delete the duplicate copyrights!! 
                if len(copyright_list)>1:
                    list(set(copyright_list) - set(copyright_list))

                
                #Remove copyright keyword!! Add the copyright(s) to third_party_contents!!
                for cr in copyright_list:
                    if cr == "Copyright" or cr == "copyright" or cr == "Copyright (C)  ":
                        continue
                    third_party_contents[title]["%s"%version]["%s"%(sheet.cell(row,component_col).value)]["%s"%(sheet.cell(row,component_col+1).value)].append(cr) 

                is_dict_license = False
                
                for license_dict in license_info:
                    
                    license_version = None
                    is_license_version_found = False
                    license_name_from_tuple = None
                    is_license_name_exist = False
                    
                    if isinstance(license_dict, tuple) is True:
                                                        
                        if any((title).find(license_name) != -1 for license_name in license_dict): #Find the license name from key tuple
                            
                            license_name_from_tuple = license_dict                                
                            is_license_name_exist = True
                            break
                        
                    elif (license_dict).find(title) != -1 or (title).find(license_dict) != -1: #Find the license name
                        license_name_from_tuple = license_dict
                        is_license_name_exist = True
                        break
                    else:
                        continue
                
                if is_license_name_exist == True:
                    
                    for v in license_info[license_name_from_tuple].keys():
                        
                        if (sheet.cell(row,license_col).value).find(v) != -1: #Find the license version
                            license_version = v
                            is_license_version_found = True
                            break
                     
                    if not is_license_version_found:
                        license_version = list(license_info[license_name_from_tuple].keys())[0]

                    
                    m = SequenceMatcher(None, web_terms,  license_info[license_name_from_tuple]["%s"%license_version])
               
                    if m.ratio() < 0.8: #Check matching ratio!!
                        third_party_contents[title]["%s"%version]["terms"] = license_info[license_name_from_tuple]["%s"%license_version]
                        is_dict_license = True
                    elif third_party_contents[title]["%s"%version]["terms"] == "":
                        third_party_contents[title]["%s"%version]["terms"] = web_terms

                else:
                    if third_party_contents[title]["%s"%version]["terms"] == "":
                        third_party_contents[title]["%s"%version]["terms"] = web_terms
                        
            
            
    #Format final third_party_contents from "None" version!
    for name in third_party_contents:
        is_del = False
        for version in third_party_contents[name]:
          
            if version == "None" and len(third_party_contents[name]) > 1:
                
                max_ver = max(val for val in third_party_contents[name] if re.match('\d+', val) != None)

                if max_ver != None:
                    del third_party_contents[name]["None"]["terms"]

                    for content in third_party_contents[name]["None"]:
                        
                        third_party_contents[name][max_ver].update(third_party_contents[name]["None"])

                        break
                    
                    try:
                        del third_party_contents[name][version]
                        is_del = True
                    except Exception:
                        pass

                    if is_del == True:
                        break
                    
            if is_del == True:
                break

                            
    #Missing copyright notice/license terms!!
    for name in third_party_contents:
        
        for version in third_party_contents[name]:

            for content in third_party_contents[name][version]:

                if content == "term" and third_party_contents[name][version][content] == None:
                    third_party_contents[name][version][content] = "Missing license terms!!"
                    
                if isinstance(third_party_contents[name][version][content], dict) == True:
                    
                    for cp in third_party_contents[name][version][content]:
                        if len(third_party_contents[name][version][content][cp]) == 0:
                            third_party_contents[name][version][content][cp] = ["Missing copyright!!"]
                            break
                            
    print(json.dumps(third_party_contents, indent=4))
    return third_party_contents
#End of the function generate_third_party_license_text.
    



#Function to create a third party license text file.
def create_third_party_license_text(third_party_contents, sheet_name, unit):

    if unit == "PMT":
        
        file_name = "Third_Party_Licenses.rtf"

        fp = open(file_name,"w")

        text = """
Third Party Licenses including Open Source Software
--------------------------------------------------------------------------------
Honeywell products use software provided by third parties, including open source software.  The following copyright statements and licenses apply to various components that are distributed with various Honeywell products.  The Honeywell product that includes this file does not necessarily use all, or any, of the third party software components referred to below.  Licensee must fully agree and comply with these license terms or must not use these components.  The third party license terms apply only to the respective software to which the license pertains, and the third party license terms do not apply to the Honeywell software.  The Honeywell software is licensed under the Honeywell End User License Agreement, which may be found in the file Honeywell_License.rtf.  If required, copies of the GPL, LGPL, and other licenses are at the end of this file in the APPENDIX.
--------------------------------------------------------------------------------
--------------------------------------------------------------------------------
Any source code that is required to be made available pursuant to its license may be found with the materials accompanying the software package or at http://www.honeywell.com/ps/thirdpartylicenses, or if the source code is not available at either location, then the applicable source code may be obtained by sending a money order or check payable to Honeywell International Inc. in the amount of US$19.99 (to cover shipping and handling) to: Honeywell International Inc., Honeywell Process Solutions, 1860 W. Rose Garden Lane, Phoenix, AZ 85027, Attn: Legal Department – Open Source. Please write “Open Source Code” and identify the Honeywell product in the memo line of your payment. The offer to obtain a copy of source code that is required to be made available pursuant to its license is valid for the shortest period of time, if any, specified by the applicable license.

"""
        fp.write("%s\n\n\n"%text)

        
    else:    

        file_name = "Open Source License %s.txt"%sheet_name
        
        fp = open(file_name,"w")
        fp.write("Licenses of Third Party Software\n ================================\n\nThis product contains software provided by third parties, which may include the below listed components. The Honeywell product that includes this file does not necessarily use all of the third party software components referred to below.  \n\n\n\n")
    
    for name in third_party_contents:

        for version in third_party_contents[name]:

            if re.search(r'license', name, re.IGNORECASE) == None:
                if version == "None":
                    fp.write("Package(s) using *** %s License\n"%name)
                else:
                    fp.write("Package(s) using *** %s License Version %s \n"%(name, version))
            else:
                if version == "None":
                    fp.write("Package(s) using *** %s\n"%name)
                else:
                    fp.write("Package(s) using *** %s Version %s \n"%(name, version))
                
            fp.write("============================\n\n")
            component_count = 0
            
            for sub_cont in third_party_contents[name][version]:
            
                if sub_cont == "terms":
                    continue

                for sub in third_party_contents[name][version][sub_cont]:
                    fp.write("%s) %s  %s\n"%(component_count+1,titlecase(sub_cont), sub))
                    component_count += 1
                
                    for notice in third_party_contents[name][version][sub_cont][sub]:
                        try:
                            fp.write("%s\n"%(notice))
                        except Exception:
                            pass
                        fp.write("\n")
                    
                    fp.write("\n")
                    
                fp.write("\n")
                 
            fp.write("\n")
            fp.write("License Text:\n")
            fp.write("--------------------------\n")
            try:
                fp.write("%s\n"%(third_party_contents[name][version]["terms"]))
            except Exception:
                pass
            fp.write("\n\n\n\n")
        
    fp.close()

    return (True, file_name)

#End of the function create_third_party_license_text.    




def backend(uploaded_file, sheet_name, unit):

    try:
        excel_file = Excel(uploaded_file, sheet_name)
    except Exception:
        raise Exception("Error loading the file!")

    #Validate file type, return file name and file type.    
    (is_excel_type,file_name,file_type) = excel_file.is_excel()

    if not is_excel_type:
        return (False, "Please upload the excel sheet!")

    #Parse the specified excel sheet name.
    err_details, sheet_details = excel_file.parse_excel_sheet()
    if err_details:
        return (False, error[err_details])
    
    #Find the component,license and row value numbers in the sheet.
    label_err, component_col, license_col, row_val =  excel_file.find_labeled_numbers(sheet_details)   
    if label_err:
        raise Exception(error[label_err])

    third_party_contents = {} #Final dictionary!
    
    third_party_contents = generate_third_party_license_text(file_type, uploaded_file, sheet_name, sheet_details, component_col, license_col, row_val, third_party_contents)

    #Function to create a third party license text.
    return create_third_party_license_text(third_party_contents, sheet_name, unit)

############################################################################END of BACK-END####################################################################






############################################################################BEGINNING of FRONT-END####################################################################

'''try:
    from Tkinter import *
except ImportError:
'''

from tkinter import *
    
from tkinter import messagebox

try:
    import ttk
    py3 = 0
except ImportError:
    import tkinter.ttk as ttk
    py3 = 1
from tkinter import filedialog

try:
    from tkcalendar import Calendar
except ImportError:
    import pip
    pip.main(['install', 'tkcalendar'])
    from tkcalendar import Calendar
    
#import billing_GUI_support



Ftype=[('Excel (*.xls*)', '*.xl*'),('Any File (*.*)', '*')]
#Beginning of Open source License Text Generation GUI function!
def olt_gui(): 
    root = Tk()
    
    #root.iconbitmap('oslt.ico')
    top = open_source_License_ui(root)
                    
    root.mainloop()
#End of olt_gui finction.



'''
you can change any thing from here, Please Be Adviced, with great power comes great responsibility.
Only change if you know what you are doing responsibly
'''

class open_source_License_ui:

    def __init__(self, top=None):
        '''This class configures and populates the toplevel window.
           top is the toplevel containing window.'''
        _bgcolor = '#e29e98'  # X11 color: 'gray85'
        _fgcolor = '#000000'  # X11 color: 'black'
        _compcolor = '#e29e98' # X11 color: 'gray85'
        _ana1color = '#e29e98' # X11 color: 'gray85' 
        _ana2color = '#e29e98' # X11 color: 'gray85'

        self.top = top

        top.geometry("600x500+307+30")
        top.title("Third Party License Text Generator!")
        top.configure(background="#e29e98")
        top.configure(highlightbackground="#e29e98")
        top.configure(highlightcolor="black")

        self.var = StringVar()
    
        self.browse = Button(top)
        self.browse.place(relx=0.80, rely=0.07, height=24, width=94)
        self.browse.configure(activebackground="#e29e98")
        self.browse.configure(activeforeground="#000000")
        self.browse.configure(background="#e29e98")
        self.browse.configure(command=self.Browse_file)
        self.browse.configure(disabledforeground="#a3a3a3")
        self.browse.configure(foreground="#000000")
        self.browse.configure(highlightbackground="#e29e98")
        self.browse.configure(highlightcolor="black")
        self.browse.configure(pady="0")
        self.browse.configure(text='''Browse''')


        self.entry_field = Entry(top)
        self.entry_field.place(relx=0.03, rely=0.07, relheight=0.04, relwidth=0.75)
        self.entry_field.configure(background="white")
        self.entry_field.configure(disabledforeground="#a3a3a3")
        self.entry_field.configure(font="TkFixedFont")
        self.entry_field.configure(foreground="#000000")
        self.entry_field.configure(highlightbackground="#e29e98")
        self.entry_field.configure(highlightcolor="black")
        self.entry_field.configure(insertbackground="black")
        self.entry_field.configure(selectbackground="#c4c4c4")
        self.entry_field.configure(selectforeground="black")


        self.Label1 = Label(top)
        self.Label1.place(relx=0.03, rely=0.02, height=21, width=544)
        self.Label1.configure(activebackground="#f9f9f9")
        self.Label1.configure(activeforeground="black")
        self.Label1.configure(anchor=W)
        self.Label1.configure(background="#e29e98")
        self.Label1.configure(disabledforeground="#a3a3a3")
        self.Label1.configure(foreground="#000000")
        self.Label1.configure(highlightbackground="#e29e98")
        self.Label1.configure(highlightcolor="black")
        self.Label1.configure(text='''Analyzed excel sheet from Code Center:''')


        self.Label7 = Label(top)
        self.Label7.place(relx=0.03, rely=0.38, height=21, width=174)
        self.Label7.configure(activebackground="#f9f9f9")
        self.Label7.configure(activeforeground="black")
        self.Label7.configure(background="#e29e98")
        self.Label7.configure(disabledforeground="#a3a3a3")
        self.Label7.configure(foreground="#000000")
        self.Label7.configure(highlightbackground="#e29e98")
        self.Label7.configure(highlightcolor="black")
        self.Label7.configure(anchor=W)
        self.Label7.configure(text='''  Excel Sheet Name:''')

        
        self.excel_sheet_name = Entry(top)
        self.excel_sheet_name.place(relx=0.04, rely=0.42, relheight=0.04, relwidth=0.23)
        self.excel_sheet_name.configure(background="white")
        self.excel_sheet_name.configure(disabledforeground="#a3a3a3")
        self.excel_sheet_name.configure(font="TkFixedFont")
        self.excel_sheet_name.configure(foreground="#000000")
        self.excel_sheet_name.configure(highlightbackground="#e29e98")
        self.excel_sheet_name.configure(highlightcolor="black")
        self.excel_sheet_name.configure(insertbackground="black")
        self.excel_sheet_name.configure(selectbackground="#c4c4c4")
        self.excel_sheet_name.configure(selectforeground="black")
        

        self.Compile_button = Button(top)
        self.Compile_button.place(relx=0.55, rely=0.85, height=30, width=165)
        self.Compile_button.configure(activebackground="#e29e98")
        self.Compile_button.configure(activeforeground="#000000")
        self.Compile_button.configure(background="#e29e98")
        self.Compile_button.configure(disabledforeground="#a3a3a3")
        self.Compile_button.configure(foreground="#000000")
        self.Compile_button.configure(highlightbackground="#e29e98")
        self.Compile_button.configure(highlightcolor="black")
        self.Compile_button.configure(pady="0")
        self.Compile_button.configure(text='''Run''')
        self.Compile_button.configure(command=self.Run)
        

        self.Label8 = Label(top)
        self.Label8.place(relx=0.55, rely=0.38, height=23, width=344)
        self.Label8.configure(activebackground="#f9f9f9")
        self.Label8.configure(activeforeground="black")
        self.Label8.configure(anchor=W)
        self.Label8.configure(background="#e29e98")
        self.Label8.configure(disabledforeground="#a3a3a3")
        self.Label8.configure(foreground="#000000")
        self.Label8.configure(highlightbackground="#e29e98")
        self.Label8.configure(highlightcolor="black")
        self.Label8.configure(text='''Third Party License Text Generator :''')

        self.Label9 = Label(top)
        self.Label9.place(relx=0.03, rely=0.93, height=21, width=394)
        self.Label9.configure(activebackground="#f9f9f9")
        self.Label9.configure(activeforeground="black")
        self.Label9.configure(anchor=NW)
        self.Label9.configure(background="#e29e98")  
        self.Label9.configure(disabledforeground="#a3a3a3")
        self.Label9.configure(foreground="#000000")
        self.Label9.configure(highlightbackground="#e29e98")
        self.Label9.configure(highlightcolor="black")
        self.Label9.configure(text='''Copyright @2018 By GALS OSS''')

            
        self.Aero_radio = Radiobutton(top)
        self.Aero_radio.place(relx=0.5695, rely=0.43, relheight=0.04, relwidth=0.09, width=-5, height=-5)
        self.Aero_radio.configure(activebackground="#e29e98")
        self.Aero_radio.configure(activeforeground="#000000")
        self.Aero_radio.configure(background="#e29e98")
        self.Aero_radio.configure(disabledforeground="#a3a3a3")
        self.Aero_radio.configure(foreground="#000000")
        self.Aero_radio.configure(highlightbackground="#e29e98")
        self.Aero_radio.configure(highlightcolor="black")
        self.Aero_radio.configure(justify=LEFT)
        self.Aero_radio.configure(offrelief="flat")
        self.Aero_radio.configure(overrelief="flat")
        self.Aero_radio.configure(text='''AERO''')
        self.Aero_radio.configure(value="AERO")
        self.Aero_radio.configure(variable=self.var)
        
        self.Homes_radio = Radiobutton(top)
        self.Homes_radio.place(relx=0.566, rely=0.48, relheight=0.04, relwidth=0.09, width=-5, height=-5)
        self.Homes_radio.configure(activebackground="#e29e98")
        self.Homes_radio.configure(activeforeground="#000000")
        self.Homes_radio.configure(background="#e29e98")
        self.Homes_radio.configure(disabledforeground="#a3a3a3")
        self.Homes_radio.configure(foreground="#000000")
        self.Homes_radio.configure(highlightbackground="#e29e98")
        self.Homes_radio.configure(highlightcolor="black")
        self.Homes_radio.configure(justify=LEFT)
        self.Homes_radio.configure(offrelief="flat")
        self.Homes_radio.configure(overrelief="flat")
        self.Homes_radio.configure(text='''HBT''')
        self.Homes_radio.configure(value="HBT")
        self.Homes_radio.configure(variable=self.var)

        self.SPS_radio = Radiobutton(top)
        self.SPS_radio.place(relx=0.564, rely=0.53, relheight=0.04, relwidth=0.09, width=-5, height=-5)
        self.SPS_radio.configure(activebackground="#e29e98")
        self.SPS_radio.configure(activeforeground="#000000")
        self.SPS_radio.configure(background="#e29e98")
        self.SPS_radio.configure(disabledforeground="#a3a3a3")
        self.SPS_radio.configure(foreground="#000000")
        self.SPS_radio.configure(highlightbackground="#e29e98")
        self.SPS_radio.configure(highlightcolor="black")
        self.SPS_radio.configure(justify=LEFT)
        self.SPS_radio.configure(offrelief="flat")
        self.SPS_radio.configure(overrelief="flat")
        self.SPS_radio.configure(text='''SPS''')
        self.SPS_radio.configure(value="SPS")
        self.SPS_radio.configure(variable=self.var)

        self.PMT_radio = Radiobutton(top)
        self.PMT_radio.place(relx=0.570, rely=0.58, relheight=0.04, relwidth=0.09, width=-5, height=-5)
        self.PMT_radio.configure(activebackground="#e29e98")
        self.PMT_radio.configure(activeforeground="#000000")
        self.PMT_radio.configure(background="#e29e98")
        self.PMT_radio.configure(disabledforeground="#a3a3a3")
        self.PMT_radio.configure(foreground="#000000")
        self.PMT_radio.configure(highlightbackground="#e29e98")
        self.PMT_radio.configure(highlightcolor="black")
        self.PMT_radio.configure(justify=LEFT)
        self.PMT_radio.configure(offrelief="flat")
        self.PMT_radio.configure(overrelief="flat")
        self.PMT_radio.configure(text='''PMT''')
        self.PMT_radio.configure(value="PMT")
        self.PMT_radio.configure(variable=self.var)

        self.BT_radio = Radiobutton(top)
        self.BT_radio.place(relx=0.5755, rely=0.63, relheight=0.04, relwidth=0.09, width=-5, height=-5)
        self.BT_radio.configure(activebackground="#e29e98")
        self.BT_radio.configure(activeforeground="#000000")
        self.BT_radio.configure(background="#e29e98")
        self.BT_radio.configure(disabledforeground="#a3a3a3")
        self.BT_radio.configure(foreground="#000000")
        self.BT_radio.configure(highlightbackground="#e29e98")
        self.BT_radio.configure(highlightcolor="black")
        self.BT_radio.configure(justify=LEFT)
        self.BT_radio.configure(offrelief="flat")
        self.BT_radio.configure(overrelief="flat")
        self.BT_radio.configure(text='''CORP''')
        self.BT_radio.configure(value="CORP")
        self.BT_radio.configure(variable=self.var)
        


    def Browse_file(self):
        K = filedialog.askopenfilename(filetypes= Ftype)
        self.entry_field.delete(0,END)
        self.entry_field.insert(0,K)
        
    
    def Run(self):

        if self.entry_field.get() == "":
            messagebox.showwarning("Warining", "Please upload the excel file!", size = 0.5)

        else:
            if self.excel_sheet_name.get() == "":
                messagebox.showwarning("Warining", "Please add the excel sheet name!")
            else:
                if self.var.get() == "":
                    messagebox.showwarning("Warining", "Please select the business unit!")
                else:

                    status, msg = backend(self.entry_field.get(), self.excel_sheet_name.get(), self.var.get())
                
                    if status == True:
                        messagebox.showinfo("Success", "%s is created successfully!!"%msg)
                        
                    else:
                        messagebox.showerror("Error", msg)

        
    

if __name__ == '__main__':
    olt_gui()
#end of entrire program



         
############################################################################END of FRONT-END####################################################################
